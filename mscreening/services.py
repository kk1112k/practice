import os
import math
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
from datetime import datetime
import io
from hangul_utils import split_syllable_char, split_syllables, join_jamos
from symspellpy import SymSpell, Verbosity
import pickle
import pandas as pd
import numpy as np
import boto3
import json
import time
import botocore
import re
import time

from velzon.settings import DJANGO_DRF_FILEPOND_FILE_STORE_PATH
from velzon.utils import BUCKET, call_csv, delete_old_files, isNullChk, s3_client


kname_key = 'kcia-korean-name/fullfile/Korean_material_name.csv'
K_names = call_csv(BUCKET, kname_key)

obj_cd = s3_client.get_object(
    Bucket=BUCKET, Key='marketing_risk/hwahae_meiri/화해_메이리_결과_코드.xlsx')
hwahae_meiri2_cd = pd.read_excel(io.BytesIO(obj_cd['Body'].read()))

cir_safety = call_csv(bucket_nm='derma-material-screening',
                      key_nm="cir-safety-data/full_file/cir_safety_full.csv")

obj = s3_client.get_object(Bucket=BUCKET, Key='china-rule-book/중국 기사용 원료.xlsx')
chinese_data = pd.read_excel(io.BytesIO(obj['Body'].read()))
chinese_data.columns = ['순번', 'Chinese_name',
                        'INCI_name', 'wash_max_perc', 'non_wash_max_perc', '비고']
chinese_data['순번'] = np.where((chinese_data['순번'].isnull() == True) & (
    chinese_data['Chinese_name'].isnull() == False), chinese_data['순번'].fillna(method='ffill'), chinese_data['순번'])

chinese_data['순번_shift'] = chinese_data['순번'].shift(1)
chinese_data['INCI_name'] = np.where((chinese_data['INCI_name'].isnull() == True) & (chinese_data['Chinese_name'].isnull() == False) & (
    chinese_data['순번'] == chinese_data['순번_shift']), chinese_data['INCI_name'].fillna(method='ffill'), chinese_data['INCI_name'])
chinese_data = chinese_data[chinese_data['순번'].isnull() == False]
chinese_data = chinese_data.drop(['순번_shift'], axis=1)


dict_1 = call_csv(bucket_nm=BUCKET,
                  key_nm="kcia-korean-name/korean_name_dict/korea_material_dict_syllables.txt", sep=" ", header=None)
dict_1.to_csv("korea_material_dict_syllables.txt",
              sep=" ", index=False, header=None)

# 분석을 위한 기초데이터 셋 2
obj_pur = s3_client.get_object(Bucket=BUCKET, Key='china-rule-book/원료목적표.xlsx')
purpose_list = pd.read_excel(io.BytesIO(obj_pur['Body'].read()))

obj_tb6 = s3_client.get_object(
    Bucket=BUCKET, Key='china-rule-book/표6착색제_사용범위.xlsx')
table_6_ = pd.read_excel(io.BytesIO(obj_tb6['Body'].read()))

obj_tb4 = s3_client.get_object(
    Bucket=BUCKET, Key='china-rule-book/표4방부제_농도.xlsx')
table_4_ = pd.read_excel(io.BytesIO(obj_tb4['Body'].read()))

obj_tb3 = s3_client.get_object(
    Bucket=BUCKET, Key='china-rule-book/표3사용제한물질_농도.xlsx')
rule_table3_limit_ = pd.read_excel(io.BytesIO(obj_tb3['Body'].read()))

obj_tb5 = s3_client.get_object(
    Bucket=BUCKET, Key='china-rule-book/표5자외선_농도.xlsx')
table_5_ = pd.read_excel(io.BytesIO(obj_tb5['Body'].read()))

gcites = s3_client.get_object(
    Bucket=BUCKET, Key="cites-data/global-cites/word/global_cites_wordlist_dict.pkl")
cites_wordlist_file = pd.read_pickle(io.BytesIO(gcites['Body'].read()))

# Get Global Cites File
gcites_d = s3_client.get_object(
    Bucket=BUCKET, Key="cites-data/global-cites/list/G_CITES_data.csv")
g_cites_data = pd.read_csv(io.BytesIO(gcites_d['Body'].read()))

ccites = s3_client.get_object(
    Bucket=BUCKET, Key="cites-data/china-cites/word/chinese_cites_wordlist_dict.pkl")
chines_cites_wordlist_file = pd.read_pickle(io.BytesIO(ccites['Body'].read()))

ccites_d = s3_client.get_object(
    Bucket=BUCKET, Key="cites-data/china-cites/list/chinese_cites_file.xlsx")
chinese_cites = pd.read_excel(io.BytesIO(ccites_d['Body'].read()))

# 1. 위생허가용 복합성분표 원본 생성


def Data_Load_process(source):
    """
    /**
    * 위생허가용 복합성분표 원본 
    * @param {string} source xlsx 파일 경로
    * @returns {DataFrame} dtfile2 위생허가용 복합성분표 원본 데이터
    * @returns {DataFrame} status 유화제여부 및 효능 체크 상태 결과 데이터
    * @returns {int} col_length 위생허가용 복합성분표 원본 컬럼 갯수

    */
    """
    def Get_product_status(source):
        """
        /**
        * 유화제여부 및 효능 체크
        * @param {string} source xlsx 파일 경로
        * @returns {string} 특수/비특수

        */
        """
        dtfile = pd.read_excel(source)
        dtlen = min(20, len(dtfile))
        for i in range(0, dtlen):
            check_point = dtfile.iloc[i, 0]
            if check_point == "제품효능":
                efficacy = dtfile.iloc[i, 1]
            if check_point == "유화제여부":
                if dtfile.iloc[i, 1] == "유화제":
                    emulse = 'True'
                else:
                    emulse = 'False'
            if check_point == "제품명":
                product_name = dtfile.iloc[i, 1]

            if check_point == "특수/비특수":
                speciality = dtfile.iloc[i, 1]

            if check_point == "제품타입":
                type_product = dtfile.iloc[i, 1]

        product_status = pd.DataFrame(columns=['제품명', '유화제여부', '제품효능'])
        try:
            product_status.loc[0, '제품명'] = product_name
        except:
            product_status.loc[0, '제품명'] = ""
        try:
            product_status.loc[0, '유화제여부'] = emulse
        except:
            product_status.loc[0, '유화제여부'] = ""
        try:
            product_status.loc[0, '제품효능'] = efficacy
        except:
            product_status.loc[0, '제품효능'] = ""

        try:
            product_status.loc[0, '특수/비특수'] = speciality
        except:
            product_status.loc[0, '특수/비특수'] = ""

        try:
            product_status.loc[0, '제품타입'] = type_product
        except:
            product_status.loc[0, '제품타입'] = ""

        def status_modify(status):
            """
            /**
            * 제품 효능에 따른 특수/비특수 데이터 변경
            * @param {DataFrame} 제품 상태 데이터
            * @returns {DataFrame} 변경된 데이터

            */
            """
            if (status.loc[0, '특수/비특수'] == "") & (status['제품효능'].str.contains('미백') == True)[0]:
                status.loc[0, '특수/비특수'] = '특수'
            elif (status.loc[0, '특수/비특수'] == "") & (status['제품효능'].str.contains('자외선차단') == True)[0]:
                status.loc[0, '특수/비특수'] = '특수'

            else:
                status.loc[0, '특수/비특수'] = '비특수'
            return (status)

        product_status = status_modify(product_status)
        return (product_status)

    # Load Data ('순번'이 있는 행을 찾아서 거기까지 skiprow를 실시하는 방식입니다)

    dtfile = pd.read_excel(source, skiprows=6)

    # 첫번재 순번이 있는 행 찾기 총 20행
    if dtfile.columns[0] != "순번":
        dtfile = pd.read_excel(source)
        count = 0
        while count < 20:
            if dtfile.iloc[count, 0] == "순번":
                break
            else:
                count = count + 1
        skprow = count + 1
        dtfile = pd.read_excel(source, skiprows=skprow)
    else:
        pass

    col_length = len(dtfile.columns)
    # Column을 바꾸기 위한 dict 제작 {추후 애러가 날 경우 칼럼 내용만 더 업데이트 하는형태면 충분}
    column_dictionary = {'순번': '순번',
                         'Korean Name ': "KoreanName",
                         'Chinese Name': "ChineseName",
                         'INCI': "INCI",
                         'RM or ingredient % \nin fla': "rm_ing_fla",
                         'RM or Ingredient % in fla': 'rm_ing_fla',
                         'ingredient % in RM': 'ingre_in_rm',
                         'Ingredient % in RM': 'ingre_in_rm',
                         'Actual Wt(%)': 'actual_wt',
                         'Ingredient function': "Ingredient_Function",
                         'Ingredient Function': "Ingredient_Function",
                         'CAS NO.': "CAS_No",
                         'Raw material trade name\n(원료명）': "Raw_t_name",
                         'Raw Material Manufacturer name\n(원료사 명칭）': "Raw_m_name",
                         'COA소지여부\n(o/x로 표기)': "COA"}

    # Change Column name
    dtfile.rename(columns=column_dictionary, inplace=True)

    # Drop blank rows
    dtfile2 = dtfile.drop(dtfile[(dtfile['순번'].isnull() == True) & (dtfile['KoreanName'].isnull() == True) & (
        dtfile['Ingredient_Function'].isnull() == True) & (dtfile['COA'].isnull() == True)].index)
    # Drop first row (meaningless row)
    dtfile2 = dtfile2[1:].reset_index(drop=True)
    dtfile2['rm_ing_fla'] = np.where(
        (dtfile2['rm_ing_fla'] == 0), np.nan, dtfile2['rm_ing_fla'])
    dtfile2['순번'] = np.where((dtfile2['순번'].isnull() == False) & (
        dtfile2['rm_ing_fla'].isnull() == True), np.nan,  dtfile2['순번'])
    # dtfile2['순번'] = np.where((dtfile2['순번'].isnull()==False) &(dtfile2['rm_ing_fla'].isnull()==0) , np.nan,  dtfile2['순번'] )
    dtfile2['순번'] = dtfile2['순번'].fillna(method="ffill")
    dtfile2['rm_ing_fla'] = dtfile2['rm_ing_fla'].fillna(method="ffill")

    status = Get_product_status(source)
    return (dtfile2, status, col_length)

# 2. DataFrame 데이터 정리


def Pre_data_cleaning(dtfile2):
    """
    /**
    * DataFrame 데이터 정리
    * 내용 : 1. 텍스트 머지 후 데이터 정리
    *           2. 'KoreanName' 컬럼이 null 인 경우 drop
    * @param {DataFrame} 원본 데이터
    * @returns {DataFrame} 정리 후 데이터

    */
    """
    cleaning_list = ['Ingredient_Function', 'Raw_t_name', 'Raw_m_name', 'COA']
    for col in cleaning_list:
        try:
            dtfile2 = Merged_cell_cleaning_text(dtfile2, col)

        except:
            pass

    other_cells_list = ['순번', 'rm_ing_fla']
    dtfile3 = dtfile2.drop(
        dtfile2[(dtfile2['KoreanName'].isnull() == True)].index)
    return (dtfile3)

# 칼럼에 따른 데이터 클리닝


def Merged_cell_cleaning_text(df, col_name):
    """
    /**
    * 칼럼에 따른 데이터 클리닝
    * 필수 조건 : 텍스트만 사용
    * 내용 : 텍스트 머지 파일들은 대부분 0 또는 nan value로 표기가 됩니다. 그렇기 때문에 택스트 먼저 한후 int 나 float을 처리합니다. 
    * @param {DataFrame} df 원본 데이터
    * @param {Any} df 컬럼명
    * @returns {DataFrame} df 정리 후 데이터

    */
    """
    df[col_name] = np.where(df[col_name] == 0, np.nan, df[col_name])
    df[col_name] = np.where(df[col_name].isnull(), np.nan, df[col_name])

    df[col_name] = df[col_name].fillna(method='ffill')
    return (df)

# 3. 스크리닝 작업


def Screening_Kor_step(df_file):
    """
    /**
    * 국문명 스크리닝 작업
    * @param {DataFrame} df 원본 데이터
    * @returns {DataFrame} 국문명스크리닝에서 통과한 자료
    * @returns {DataFrame} 국문명스크리닝에서 통과되지 못한 자료

    */
    """
    print("-----1-1차 스크리닝 시작")
    df2, df_nonpass = screening1_1(df_file, K_names)
    print("-----1-1차 스크리닝 완료")
    base = df2[df2['성분명'].isnull() == False]
    if len(df_nonpass) > 0:
        print("-----1-2차 스크리닝 시작")
        df_nonpass_check, df_non_passing_file2 = screening1_2(
            df_nonpass, K_names)
        add_file = df_nonpass_check[df_nonpass_check['성분명'].isnull() == False]
        base = pd.concat([base, add_file])
        print("-----1-2차 스크리닝 완료")
        if len(df_non_passing_file2) > 0:
            print("-----1-3차 스크리닝 시작")
            last, last_nonpassing = screening1_3(df_non_passing_file2, K_names)
            lastfile = last[last['성분명'].isnull() == False]
            base = pd.concat([base, lastfile])
            base = pd.concat([base, last_nonpassing])
            base['영문명'] = np.where(base['영문명'].isnull()
                                   == True, '정보없음', base['영문명'])
            base['성분명'] = np.where(base['성분명'].isnull()
                                   == True, '정보없음', base['성분명'])

            print("-----1-3차 스크리닝 완료")
            base = base.sort_index(ascending=True)
        else:
            last_nonpassing = df_non_passing_file2

    else:
        last_nonpassing = df_nonpass

    return (base, last_nonpassing)


def screening1_1(df_file, K_names):
    """
    /**
    * 불러온자료와, 국문 화장품 명칭을 매칭하여, 영문명, cas no 가져옵니다. 
    * 성분명 병합
    * @param {DataFrame} df 원본 데이터
    * @param {col_name} col_name 컬럼명
    * @returns {DataFrame} 국문명칭 전체 합쳐진 자료 1개
    * @returns {DataFrame} 합쳐지지 않은 자료 세트 1개

    */
    """
    df_file['KoreanName'] = df_file['KoreanName'].str.strip()
    K_names['성분명'] = K_names['성분명'].str.strip()
    K_names['구명칭'] = K_names['구명칭'].str.strip()
    df_file2 = pd.merge(df_file, K_names, how='left',
                        left_on='KoreanName', right_on='성분명')
    df_non_passing_file = df_file2[df_file2['성분명'].isnull()]
    df_file2['Kor_Comment'] = ""
    print("1차 국문명스크리닝에서 통과되지 못한 자료들은 : 총 {} 건".format(len(df_non_passing_file)))
    return (df_file2, df_non_passing_file)


def screening1_2(df_nonpass, K_names):
    """
    /**
    * 불러온자료와, 국문 화장품 명칭을 매칭하여, 영문명, cas no 가져옵니다. 
    * '성분코드', '성분명', '영문명', 'CAS No', '구명칭' 컬럼 제거 후 구명칭 병합
    * @param {DataFrame} df 원본 데이터
    * @param {col_name} col_name 컬럼명
    * @returns {DataFrame} 국문명칭 전체 합쳐진 자료 1개
    * @returns {DataFrame} 합쳐지지 않은 자료 세트 1개

    */
    """
    df_nonpass = df_nonpass.drop(
        ['성분코드', '성분명', '영문명', 'CAS No', '구명칭'], axis=1)
    df_nonpass['KoreanName'] = df_nonpass['KoreanName'].str.strip()
    K_names['구명칭'] = K_names['구명칭'].str.strip()
    df_file2 = pd.merge(df_nonpass, K_names, how='left',
                        left_on='KoreanName', right_on='구명칭')
    df_non_passing_file = df_file2[df_file2['성분명'].isnull()]
    print("1-2차 구명칭 국문명스크리닝에서 통과되지 못한 자료들은 : 총 {} 건".format(len(df_non_passing_file)))
    df_file2['Kor_Comment'] = "국문명 명칭이 구명칭 입니다."
    return (df_file2, df_non_passing_file)


def screening1_3(df_non_passing_file2, K_names):
    """
    오탈자 변화시켜서 매칭합니다. 
    """
    SC = spell_check()
    df_non_passing_file2 = df_non_passing_file2.drop(
        ['성분코드', '성분명', '영문명', 'CAS No', '구명칭'], axis=1)
    df_non_passing_file2['adj_KN'] = df_non_passing_file2['KoreanName'].apply(
        lambda x: SC.spell_check_suggestion(x, rank=1))
    df_non_passing_file3 = pd.merge(
        df_non_passing_file2,  K_names, how='left', left_on='adj_KN', right_on='성분명')
    df_non_passing_file3['Kor_Comment'] = "오탈자가 존재합니다. 수정어는 : " + \
        df_non_passing_file3['adj_KN']
    df_non_passing_file3 = df_non_passing_file3.drop(['adj_KN'], axis=1)
    df_non_passing_file3 = df_non_passing_file3.set_index(
        df_non_passing_file2.index)
    df_non_passing_file = df_non_passing_file3[df_non_passing_file3['성분명'].isnull(
    )]

    if len(df_non_passing_file) > 0:
        # 구명칭도 오류가 있을수 있다. 구명칭까지 같이 하자
        try:
            # 명칭 애러가 있으면 머지에서 에러가 남
            df_non_passing_file = df_non_passing_file.drop(
                ['성분코드', '성분명', '영문명', 'CAS No', '구명칭'], axis=1)
            print(df_non_passing_file.columns)
            df_non_passing_file3_1 = pd.merge(
                df_non_passing_file,  K_names, how='left', left_on='adj_KN', right_on='구명칭')
            df_non_passing_file3_1['Kor_Comment'] = "오탈자가 존재합니다.(게다가 구명칭입니다.) 수정어는 : " + \
                df_non_passing_file3['adj_KN']
            df_non_passing_file4 = pd.concat(
                [df_non_passing_file3, df_non_passing_file3_1])
            df_non_passing_file = df_non_passing_file3[df_non_passing_file3['성분명'].isnull(
            )]
        except:
            print('통과못해서 그냥 내보냅니다')
            print(df_non_passing_file)
            df_non_passing_file4 = df_non_passing_file3

    else:
        df_non_passing_file4 = df_non_passing_file3

    print("1-3차 오탈자 국문명스크리닝에서 통과되지 못한 자료들은 : 총 {} 건".format(len(df_non_passing_file)))
    return (df_non_passing_file4, df_non_passing_file)


class spell_check():
    """
    국문의 spelling check를 도와주는 function
    먼저 클래스를 불러온 후 
    spell check suggestion을 사용합니다. 
    구분한 이유 : symspell starting이 느립니다. 그렇기 때문에 한번만 돌리고 그다음은 계속 사용하는 방식으로 하면 빠르게 사용가능합니다. 

    """

    def __init__(self, dict_path=None):
        """
        시작구문입니다. 
        시작과 동시에 dictionary 를 로드 합니다. 
        dictionary가 있다면, path를 추가하면 됩니다. 
        기본으로 제공되는 dictionary는 한국 화장품 성분사전입니다. 
        """
        self.sym_spell = SymSpell(
            max_dictionary_edit_distance=7, prefix_length=8)
        self.Join_jamos = join_jamos
        if dict_path == None:
            diction_path = 'korea_material_dict_syllables.txt'
        else:
            diction_path = dict_path
        # dictionary call
        self.sym_spell.load_dictionary(diction_path, 0, 1, " ", 'UTF-8')

    def spell_check_suggestion(self, term, rank):
        """
        LOGIC : 가장 적게 틀린 문구를 찾는것입니다. 
        PARAM term: 은 문자열 str입니다. 해당문자열을 분해한 후 매칭합니다. 
        PARAM rank: 몇개를 가져올지 숫자로 씁니다. 1~3까지입니다.
        RETURN : RANK개 만큼 가져온 후보 군입니다. 
        """
        # check spell
        term = split_syllables(term)
        # output results
        suggestions = self.sym_spell.lookup(term, Verbosity.ALL)
        # show top 3 candidates
        try:
            for sugg in suggestions[:3]:
                print(sugg.term, self.Join_jamos(
                    sugg.term), sugg.distance, sugg.count)
            if rank == 1:
                return (self.Join_jamos(suggestions[0].term))
            elif rank == 2:
                return (self.Join_jamos(suggestions[0].term), self.Join_jamos(suggestions[1].term))
            elif rank == 3:
                return (self.Join_jamos(suggestions[0].term), self.Join_jamos(suggestions[1].term), self.Join_jamos(suggestions[2].term))
        except:
            return ("국문명 없음")

# 4. 중문 매칭 프로세스


def Chinese_matching_Process(df, chinese_data):
    """
    이 펑션은 중문 매칭 프로세스입니다.
    LOGIC : DF와 중문 명칭이 있는 기사용목록표(CHINESE DATA)를 INCI NAME기준으로 합칩니다. 
    PARAMS df : 국문명칭이 합쳐진 자료
    PARAMS chinese_data : 기사용 목록표 중 칼럼이 정리된 자료 
    RETURN : 중문과 합쳐진 결과가 나옵니다. 만약 exist 칼럼이 없으면 인키네임으로 합쳐지지 않은것. 그런경우는 사용 못합니다. 
    """
    df['inci_mat'] = df['영문명'].str.upper()
    df['inci_mat'] = df['inci_mat'].apply(str)

    df['inci_mat'] = df['inci_mat'].apply(lambda x: re.sub(r"\s", "", x))
    chinese_data['exist'] = 1
    chinese_data['inci_mat'] = chinese_data['INCI_name'].str.upper()
    chinese_data['inci_mat'] = chinese_data['inci_mat'].apply(str)
    chinese_data['inci_mat'] = chinese_data['inci_mat'].apply(
        lambda x: re.sub(r"\s", "", x))
    df_china = pd.merge(df, chinese_data.drop(
        ['순번'], axis=1), how='left', left_on='inci_mat', right_on='inci_mat')
    df_china['CN_Comment'] = np.where(
        df_china['exist'] == 1, "", "중문자료와 INCI name이 안맞습니다. 확인해야합니다.")
    return (df_china)

# 5.


def weight_check(df_china_sort):
    print("-----개별원료의 총합이 100%인지 확인합니다")
    mat_sum = df_china_sort.groupby(['순번']).head(1)['rm_ing_fla'].sum()
    if math.isclose(mat_sum, 100):
        df_china_sort['material_sum'] = 'check'
        print(f'현재 총합은 {mat_sum}')
    else:
        df_china_sort['material_sum'] = f'개별원료 총합이 100이 아닙니다. 확인부탁드립니다. 현재 총합은 {mat_sum}'
        print(f'현재 총합은 {mat_sum}')

    print("-----각 복합성분에 대해 개별 복합성분의 성분비 총합이 100% 인지 확인합니다.")
    df_china_sort['ingre_in_rm'] = df_china_sort['ingre_in_rm'].apply(float)
    _check = df_china_sort.groupby(['순번'])['ingre_in_rm'].sum().reset_index()
    _check.rename(columns={'ingre_in_rm': 'ingre_in_rm_sum'}, inplace=True)
    df_china_sort2 = pd.merge(df_china_sort, _check, how='left', on='순번')
    for i in range(0, len(df_china_sort2)):
        df_china_sort2.loc[i, 'material_in_sum'] = math.isclose(
            df_china_sort2.loc[i, 'ingre_in_rm_sum'], 100)
    df_china_sort2['material_in_sum'] = np.where(
        df_china_sort2['material_in_sum'] == True, "check", "총합이 100이 아닙니다.")

    print("-----마지막으로 실제 성분의 합이 100인지 체크합니다.")
    df_china_sort2['actual_wt_valid'] = (
        df_china_sort2['rm_ing_fla'] * df_china_sort2['ingre_in_rm'])/100
    for i in range(0, len(df_china_sort2)):
        df_china_sort2.loc[i, 'actual_wt_valid_check'] = math.isclose(
            df_china_sort2.loc[i, 'actual_wt'], df_china_sort2.loc[i, 'actual_wt_valid'])
    df_china_sort2['actual_wt_valid_check'] = np.where(
        df_china_sort2['actual_wt_valid_check'] == True, 'Check', '에러! 실제 웨이트에 대한 계산을 다시하시오')
    print("-----2차 분석 완료!.")
    return (df_china_sort2)

# 6.


cosing_annexfile = call_csv(
    bucket_nm='derma-material-screening', key_nm="ci-code/eu_ci.csv")
ci_datas2 = call_csv(bucket_nm='derma-material-screening',
                     key_nm="ci-code/CI_china_updated.csv")


def Third_round_screening_cites(df, purpose_list, table_6_, table_4_, rule_table3_limit_, table_5_, cites_wordlist_file, g_cites_data, chines_cites_wordlist_file, chinese_cites):

    # 기술규범 판단
    def Not_use_ingred_china(df):
        df['not_use_code'] = np.where(
            df['비고'].str.contains("사용금지") == True, 1, 0)
        return (df)

    # 화장품 사용 제한 여부 체크

    def third_table3_stage(df_china_sort_sum2, warning_full_dt):
        warning_full_dt['표3화장품사용제한여부'] = 'check'
        df['actual_wt_valid_gr'] = df.groupby(
            ['성분명'])['actual_wt_valid'].transform(sum)
        # 순번 영문명칭 인키네임은 따로 필요가 없음.
        warning_full_dt = warning_full_dt.drop(['순번', '영문명칭', 'INCI'], axis=1)
        # 여기는 데이터 머지하는것
        dt = pd.merge(df_china_sort_sum2, warning_full_dt,
                      left_on='Chinese_name', right_on='중문명칭', how='left')

        # 머지하고 나서 비교해야하는데, 여기서 문제가 있음.
        # 머지 후 비교시 NaN value와 빈칸이 공존하면, condition 계산시, float, str둘다 나와서 계산이 안됨
        # 둘 데이터 타입을 str으로 통일한 후 비교해주자
        for q in range(0, 10):
            var_c = q + 1
            varn = f'최대사용농도_{var_c}번째_기준'
            newvarn = f'농도초과_{var_c}번째_기준'
            dt[varn] = pd.to_numeric(dt[varn])
            # dt[varn] = dt[varn].apply(str)
            dt['actual_wt_valid_gr'] = pd.to_numeric(dt['actual_wt_valid_gr'])
            # dt['actual_wt_valid_gr'] = dt['actual_wt_valid_gr'].apply(str)
            dt[newvarn] = ""
            dt[newvarn] = np.where(
                (dt[varn] < dt['actual_wt_valid_gr']), f'{varn} 초과! 확인 필요합니다', dt[newvarn])
            dt[newvarn] = np.where((dt[varn].isnull() == True) | (
                dt[varn] == 'nan') | (dt[varn] == ''), '', dt[newvarn])

        dt['표3화장품사용제한_농도초과여부'] = ""

        # 여기서는 각 농도초과 요인들 다 합쳐주기
        for q in range(0, 10):
            var_c = q + 1
            newvarn = f'농도초과_{var_c}번째_기준'
            dt['표3화장품사용제한_농도초과여부'] = dt['표3화장품사용제한_농도초과여부'] + " " + dt[newvarn]

        dt['표3화장품사용제한_농도초과여부'] = dt['표3화장품사용제한_농도초과여부'].str.strip()
        # 데이터 중 불필요한 정보들 제거하기
        dt = dt.drop(['중문명칭', '최대사용제한_사용제한',
                      '최대사용농도_1번째_기준', '최대사용농도_2번째_기준', '최대사용농도_3번째_기준', '최대사용농도_4번째_기준',
                      '최대사용농도_5번째_기준', '최대사용농도_6번째_기준', '최대사용농도_7번째_기준', '최대사용농도_8번째_기준',
                      '최대사용농도_9번째_기준', '최대사용농도_10번째_기준', '농도초과_1번째_기준',
                      '농도초과_2번째_기준', '농도초과_3번째_기준', '농도초과_4번째_기준', '농도초과_5번째_기준',
                      '농도초과_6번째_기준', '농도초과_7번째_기준', '농도초과_8번째_기준', '농도초과_9번째_기준',
                      '농도초과_10번째_기준'], axis=1)

        dt['표3화장품사용제한여부'] = np.where(
            dt['표3화장품사용제한여부'] == "check", '이 원료는 화장품사용제한 정보를 확인해야합니다.', '')
        return (dt)

    # 준용 방부제 여부 체크

    def third_table4_stage(df_china_sort_sum2, warning_full_dt):
        warning_full_dt['표4화장품준용방부제여부'] = 'check'
        warning_full_dt = warning_full_dt.drop(['순번', '영문명칭', 'INCI'], axis=1)
        dt = pd.merge(df_china_sort_sum2, warning_full_dt,
                      left_on='Chinese_name', right_on='중문명칭', how='left')

        # 머지하고 나서 비교해야하는데, 여기서 문제가 있음.
        # 머지 후 비교시 NaN value와 빈칸이 공존하면, condition 계산시, float, str둘다 나와서 계산이 안됨
        # 둘 데이터 타입을 str으로 통일한 후 비교해주자
        for q in range(0, 10):
            var_c = q + 1
            varn = f'최대사용농도_{var_c}번째_기준'
            newvarn = f'농도초과_{var_c}번째_기준'
            dt[varn] = pd.to_numeric(dt[varn])
            # dt[varn] = dt[varn].apply(str)
            dt['actual_wt_valid_gr'] = pd.to_numeric(dt['actual_wt_valid_gr'])
            # dt['actual_wt_valid_gr'] = dt['actual_wt_valid_gr'].apply(str)
            dt[newvarn] = ""
            dt[newvarn] = np.where(
                (dt[varn] < dt['actual_wt_valid_gr']), f'{varn} 초과! 확인 필요합니다', dt[newvarn])
            dt[newvarn] = np.where((dt[varn].isnull() == True) | (
                dt[varn] == 'nan') | (dt[varn] == ''), '', dt[newvarn])

        dt['표4화장품준용방부제여부_농도초과여부'] = ""

        # 여기서는 각 농도초과 요인들 다 합쳐주기
        for q in range(0, 10):
            var_c = q + 1
            newvarn = f'농도초과_{var_c}번째_기준'
            dt['표4화장품준용방부제여부_농도초과여부'] = dt['표4화장품준용방부제여부_농도초과여부'] + " " + dt[newvarn]

        dt['표4화장품준용방부제여부_농도초과여부'] = dt['표4화장품준용방부제여부_농도초과여부'].str.strip()
        dt = dt.drop('중문명칭', axis=1)
        dt = dt.drop(['최대사용제한_사용제한',
                      '최대사용농도_1번째_기준', '최대사용농도_2번째_기준', '최대사용농도_3번째_기준', '최대사용농도_4번째_기준',
                      '최대사용농도_5번째_기준', '최대사용농도_6번째_기준', '최대사용농도_7번째_기준', '최대사용농도_8번째_기준',
                      '최대사용농도_9번째_기준', '최대사용농도_10번째_기준', '농도초과_1번째_기준',
                      '농도초과_2번째_기준', '농도초과_3번째_기준', '농도초과_4번째_기준', '농도초과_5번째_기준',
                      '농도초과_6번째_기준', '농도초과_7번째_기준', '농도초과_8번째_기준', '농도초과_9번째_기준',
                      '농도초과_10번째_기준'], axis=1)
        dt['표4화장품준용방부제여부'] = np.where(
            dt['표4화장품준용방부제여부'] == "check", '이 원료는 준용방부제 정보를 확인해야합니다.', '')

        if len(dt[dt['표4화장품준용방부제여부'] != ""]) == 0:
            dt['방부제챌린지테스트'] = "필요"
        else:
            dt['방부제챌린지테스트'] = "불필요"
        return (dt)

    # 준용 자외선
    def third_table5_stage(df_china_sort_sum2, warning_full_dt):
        warning_full_dt['표5화장품준용자외선여부'] = 'check'
        warning_full_dt = warning_full_dt.drop(['순번', '영문명칭', 'INCI'], axis=1)
        dt = pd.merge(df_china_sort_sum2, warning_full_dt,
                      left_on='Chinese_name', right_on='중문명칭', how='left')
        for q in range(0, 10):
            var_c = q + 1
            varn = f'최대사용농도_{var_c}번째_기준'
            newvarn = f'농도초과_{var_c}번째_기준'
            dt[varn] = pd.to_numeric(dt[varn])
            # dt[varn] = dt[varn].apply(str)
            dt['actual_wt_valid_gr'] = pd.to_numeric(dt['actual_wt_valid_gr'])
            # dt['actual_wt_valid_gr'] = dt['actual_wt_valid_gr'].apply(str)
            dt[newvarn] = ""
            dt[newvarn] = np.where(
                (dt[varn] < dt['actual_wt_valid_gr']), f'{varn} 초과! 확인 필요합니다', dt[newvarn])
            dt[newvarn] = np.where((dt[varn].isnull() == True) | (
                dt[varn] == 'nan') | (dt[varn] == ''), '', dt[newvarn])

        dt['표5화장품준용자외선여부_농도초과여부'] = ""
        for q in range(0, 10):
            var_c = q + 1
            newvarn = f'농도초과_{var_c}번째_기준'
            dt['표5화장품준용자외선여부_농도초과여부'] = dt['표5화장품준용자외선여부_농도초과여부'] + " " + dt[newvarn]

        dt['표5화장품준용자외선여부_농도초과여부'] = dt['표5화장품준용자외선여부_농도초과여부'].str.strip()

        dt = dt.drop(['중문명칭', '최대사용제한_사용제한', '순번_shft1',
                      '최대사용농도_1번째_기준', '최대사용농도_2번째_기준', '최대사용농도_3번째_기준', '최대사용농도_4번째_기준',
                      '최대사용농도_5번째_기준', '최대사용농도_6번째_기준', '최대사용농도_7번째_기준', '최대사용농도_8번째_기준',
                      '최대사용농도_9번째_기준', '최대사용농도_10번째_기준', '농도초과_1번째_기준',
                      '농도초과_2번째_기준', '농도초과_3번째_기준', '농도초과_4번째_기준', '농도초과_5번째_기준',
                      '농도초과_6번째_기준', '농도초과_7번째_기준', '농도초과_8번째_기준', '농도초과_9번째_기준',
                      '농도초과_10번째_기준'], axis=1)
        dt['표5화장품준용자외선여부'] = np.where(
            dt['표5화장품준용자외선여부'] == "check", '이 원료는 준용 자외선차단제 정보를 확인해야합니다.', '')
        return (dt)

    # 준용 착색제
    # CI네임을 가져옵니다. 그다음에 테스트
    def Get_CI_name_data(df, cosing_annexfile, ci_datas2):
        # cosing_annexfile = pd.read_csv('/home/ec2-user/environment/efs/DATA/MATERIAL_SCREENING/01. DATA/eu_ci.csv')
        cosing_annexfile.rename(
            columns={'Chemical name': '영문명', 'CI_name': 'CI_name_inci'}, inplace=True)
        # 여기서 합쳐진 파일은 중문명칭에 만약 없으면, CI가 INCIname으로 가야한다
        cinames_file = pd.merge(df, cosing_annexfile[[
                                '영문명', 'CI_name_inci']], left_on='영문명', right_on='영문명', how='left')

        # 중문 CI명칭 가져오기
        # ci_datas2= pd.read_csv('/home/ec2-user/environment/efs/DATA/MATERIAL_SCREENING/01. DATA/CI_china_updated.csv')
        ci_datas2.rename(
            columns={'Chinese_name': 'CI_name2', 'INCI_name': '영문명'}, inplace=True)

        # 중문 CI명칭 합하기
        cinames_file2 = pd.merge(cinames_file, ci_datas2[[
                                 'CI_name2', '영문명']], left_on="영문명", right_on='영문명', how='left')

        # 중국명칭 체크하고 없애기
        def Change_chinese_name_ciname(df):
            df2 = df.copy()
            df2['CI_Check'] = ""
            df2['CI_Check'] = df2.loc[(df2['CI_name2'].isnull() == False) & (df2['CI_name_inci'].isnull() == False), 'CI_name2'].apply(
                lambda x: f"중국기사용 원료중 착색제 코드로 들어가있는 원료입니다. 사용용도를 확인후착 중문명과 INCI명을 {x} (으)로 해주세요 그리고 이 제품은 중문명이 {x} 입니다")
            df2['CI_Check'] = np.where((df2['CI_name2'].isnull() == True) & (df2['CI_name_inci'].isnull() == False),  df2['CI_name_inci'].apply(
                lambda x: f"중국기사용 원료 착색제 코드에 포함되어있지 않습니다. 사용용도를 확인후, 착색제일경우 INCI명을 {x} (으)로 해주세요"), df2['CI_Check'])
            df2['CI_Check'] = np.where((df2['CI_name2'].isnull() == False) & (df2['CI_name_inci'].isnull() == True),  df2['CI_name2'].apply(
                lambda x: f"중국기사용 원료중 착색제 코드로 들어가있는 원료입니다. 사용용도를 확인후 중문명을 {x} (으)로 해주세요"), df2['CI_Check'])

            # df2['CI_Check']  = df2.loc[(df2['CI_name2'].isnull() == True) & (df2['CI_name_inci'].isnull()== False) , 'CI_name_inci'].apply(lambda x: f"중국기사용 원료 착색제 코드에 포함되어있지 않습니다. 사용용도를 확인후, 착색제일경우 INCI명을 {x} (으)로 해주세요")
            # df2['CI_Check']  = df2.loc[(df2['CI_name2'].isnull() == False) & (df2['CI_name_inci'].isnull()== True) , 'CI_name2'].apply(lambda x: f"중국기사용 원료중 착색제 코드로 들어가있는 원료입니다. 사용용도를 확인후 중문명을 {x} (으)로 해주세요")
            df2['CI_name'] = ""
            df2.loc[(df2['CI_name2'].isnull() == False) & (
                df2['CI_name_inci'].isnull() == True), 'CI_name'] = df2['CI_name2']
            df2.loc[(df2['CI_name2'].isnull() == True) & (
                df2['CI_name_inci'].isnull() == False), 'CI_name'] = df2['CI_name_inci']
            df2.loc[(df2['CI_name2'] == df2['CI_name_inci']),
                    'CI_name'] = df2['CI_name2']
            df2.loc[(df2['CI_name2'].isnull() == False) & (df2['CI_name_inci'].isnull(
            ) == False) & (df2['CI_name'].isnull() == False), 'CI_name'] = df2['CI_name2']

            return (df2)

        cinames_file3 = Change_chinese_name_ciname(cinames_file2)

        return (cinames_file3)

    def third_table6_stage(df, df2):
        df2['준용착색제 존재여부'] = 1
        df3 = pd.merge(df, df2[df2['color_index'] != ""][['color_index', '기타제한조건 및 요구사항_착색제', '주의사항_착색제', '사용범위_착색제', 'color', '준용착색제 존재여부', '최대사용농도_1번째_기준', '최대사용농도_2번째_기준', '최대사용농도_3번째_기준', '최대사용농도_4번째_기준',
                                                          '최대사용농도_5번째_기준', '최대사용농도_6번째_기준', '최대사용농도_7번째_기준', '최대사용농도_8번째_기준',
                                                          '최대사용농도_9번째_기준', '최대사용농도_10번째_기준']], how='left', left_on='CI_name', right_on='color_index')
        df3['준용착색제 존재여부'] = np.where(df3['준용착색제 존재여부'] == 1, 1, 0)
        dt = df3.copy()
        # 머지하고 나서 비교해야하는데, 여기서 문제가 있음.
        # 머지 후 비교시 NaN value와 빈칸이 공존하면, condition 계산시, float, str둘다 나와서 계산이 안됨
        # 둘 데이터 타입을 str으로 통일한 후 비교해주자

        for q in range(0, 10):
            var_c = q + 1
            varn = f'최대사용농도_{var_c}번째_기준'
            newvarn = f'농도초과_{var_c}번째_기준'
            dt[varn] = pd.to_numeric(dt[varn])
            # dt[varn] = dt[varn].apply(str)
            dt['actual_wt_valid_gr'] = pd.to_numeric(dt['actual_wt_valid_gr'])
            # dt['actual_wt_valid_gr'] = dt['actual_wt_valid_gr'].apply(str)
            dt[newvarn] = ""
            dt[newvarn] = np.where(
                (dt[varn] < dt['actual_wt_valid_gr']), f'{varn} 초과! 확인 필요합니다', dt[newvarn])
            dt[newvarn] = np.where((dt[varn].isnull() == True) | (
                dt[varn] == 'nan') | (dt[varn] == ''), '', dt[newvarn])

        dt['표6준용착색제여부_농도초과여부'] = ""

        # 여기서는 각 농도초과 요인들 다 합쳐주기
        for q in range(0, 10):
            var_c = q + 1
            newvarn = f'농도초과_{var_c}번째_기준'
            dt['표6준용착색제여부_농도초과여부'] = dt['표6준용착색제여부_농도초과여부'] + " " + dt[newvarn]

        dt['표6준용착색제여부_농도초과여부'] = dt['표6준용착색제여부_농도초과여부'].str.strip()
        dt = dt.drop(['최대사용농도_1번째_기준', '최대사용농도_2번째_기준', '최대사용농도_3번째_기준', '최대사용농도_4번째_기준',
                      '최대사용농도_5번째_기준', '최대사용농도_6번째_기준', '최대사용농도_7번째_기준', '최대사용농도_8번째_기준',
                      '최대사용농도_9번째_기준', '최대사용농도_10번째_기준', '농도초과_1번째_기준',
                      '농도초과_2번째_기준', '농도초과_3번째_기준', '농도초과_4번째_기준', '농도초과_5번째_기준',
                      '농도초과_6번째_기준', '농도초과_7번째_기준', '농도초과_8번째_기준', '농도초과_9번째_기준',
                      '농도초과_10번째_기준'], axis=1)

        return (dt)

    # 원료사용목적 리스트
    # 성분목적 수 체크

    def Find_Issue_U_Purpose(third_file_4_2_):
        third_file_4_2_['unique_purpose'] = third_file_4_2_.groupby(
            ['순번'])['Ingredient_Function'].transform('nunique')
        third_file_4_2_['성분목적 이슈'] = ""
        third_file_4_2_.loc[third_file_4_2_[
            'unique_purpose'] > 1, '성분목적 이슈'] = "성분목적이 1개 이상입니다"

        return (third_file_4_2_)

    def Ingredient_function_purpose_Check_(third_file_4_2_, purpose_list):
        third_file_4_2_['Ingredient_Function'] = third_file_4_2_[
            'Ingredient_Function'].fillna("")
        third_file_4_2_['Ingredient_Function lw'] = third_file_4_2_[
            'Ingredient_Function'].str.strip()
        third_file_4_2_['Ingredient_Function lw'] = third_file_4_2_[
            'Ingredient_Function lw'].str.lower()
        third_file_5_1_ = pd.merge(third_file_4_2_, purpose_list,
                                   left_on='Ingredient_Function lw', right_on='대응 영문 Function lw', how='left')
        third_file_5_1_['사용목적 존재 여부 이슈'] = ""
        third_file_5_1_.loc[third_file_5_1_['대응 영문 Function lw'].isnull(
        ) == True, '사용목적 존재 여부 이슈'] = "대응되는 영문 사용목적명이 없습니다. 적절한 사용목적인지 체크해주세요"
        return (third_file_5_1_)

    # 사용가능 제품효능만들기

    def Find_Usable_effect(third_file_5_1_):
        third_file_5_1_['사용가능 제품 효능'] = ""
        if 'Whitening Cosmetics' in third_file_5_1_['대응 영문 Function'].unique():
            third_file_5_1_['사용가능 제품 효능'] = third_file_5_1_[
                '사용가능 제품 효능'] + "미백"
        if 'Sunscreen Agent' in third_file_5_1_['대응 영문 Function'].unique():
            third_file_5_1_['사용가능 제품 효능'] = third_file_5_1_[
                '사용가능 제품 효능'] + " 자외선 차단"
        if 'Cleansing Agent' in third_file_5_1_['대응 영문 Function'].unique():
            third_file_5_1_['사용가능 제품 효능'] = third_file_5_1_[
                '사용가능 제품 효능'] + " 청결(클렌저)"
        if 'Exfoliant' in third_file_5_1_['대응 영문 Function'].unique():
            third_file_5_1_['사용가능 제품 효능'] = third_file_5_1_[
                '사용가능 제품 효능'] + " 각질제거"
        if 'Abrasive' in third_file_5_1_['대응 영문 Function'].unique():
            third_file_5_1_['사용가능 제품 효능'] = third_file_5_1_[
                '사용가능 제품 효능'] + " 각질제거"
        if 'Film Forming' in third_file_5_1_['대응 영문 Function'].unique():
            third_file_5_1_['사용가능 제품 효능'] = third_file_5_1_[
                '사용가능 제품 효능'] + " 피막형성제"
        return (third_file_5_1_)

    # 컬러런트 있으면 CI명칭을 chinese 명 바꾸기,

    def Change_CI_code_bfunction(df):
        df.loc[(df['대응 영문 Function'] == "Colorant"),
               'INCI_name'] = df['CI_name']
        df.loc[(df['대응 영문 Function'] == "Colorant") & (
            df['CI_name2'].isnull == False), 'Chinese_name'] = df['CI_name']
        return (df)

    # 특수사용목적 관련 이슈있으면 이야기 하기

    def Change_purpose_issue_update(df):
        df['특수사용목적 관련 이슈'] = ""
        df.loc[(df['대응 영문 Function'] == "Skin-Conditioning Agent"),
               '특수사용목적 관련 이슈'] = 'skin conditioning agent 존재'
        df.loc[(df['Ingredient_Function'].str.lower().str.strip(
        ) == "skin conditioning agent"), '특수사용목적 관련 이슈'] = 'skin conditioning agent 존재'

        df['안전규범 사용목적 차이 이슈'] = ""
        df.loc[(df['대응 영문 Function'] == "Preservative") & (
            df['표4화장품준용방부제여부'] == ""), '안전규범 사용목적 차이 이슈'] = '중국의 준용 방부제가 아닙니다.'
        df.loc[(df['대응 영문 Function'] == "Colorant") & (
            df['준용착색제 존재여부'] == 0), '안전규범 사용목적 차이 이슈'] = '중국의 준용 착색제가 아닙니다.'

        df.loc[(df['대응 영문 Function'] == "Sunscreen Agent") & (
            df['표5화장품준용자외선여부'] == ""), '안전규범 사용목적 차이 이슈'] = '중국의 준용 자외선차단제가 아닙니다.'

        return (df)


#     ### 글로벌 사이테스 펑션 실시
#     #Get Global Cites Word List

#     with open('/home/ec2-user/environment/efs/DATA/MATERIAL_SCREENING/01. DATA/global_cites_wordlist_dict.pkl', 'rb') as f:
#         cites_wordlist_file = pickle.load(f)

#     # Get Global Cites File

#     g_cites_data = pd.read_csv('/home/ec2-user/environment/efs/DATA/MATERIAL_SCREENING/01. DATA/G_CITES_data.csv')

    def Checking_Global_CITES(df, cites_wordlist_file, g_cites_data):
        # 변수 안에 있는 단어의 리스트가 맞는지 체크 하는것 있는것만 가져와라는 형식입니다.
        pat = '|'.join(r"\b{}\b".format(x) for x in cites_wordlist_file)

        # 매칭시 있을 수 있는 단어들의 대소문자 차이를 없애기 위해 모두 소문자로 처리
        df['cites_match_word'] = df["영문명"].str.lower(
        ).str.findall(pat).apply(" ".join)

        # 변수 생성
        df2 = pd.merge(df, g_cites_data, on='cites_match_word', how='left')

        # 경고문구 제작
        df2["Global_CITES_경고"] = ''
        df2["Global_CITES_경고"] = np.where(
            df2['cites_match_word'] != "",  'CITES 확인이 필요합니다',  df2["Global_CITES_경고"])
        df2["Global_CITES_경고"] = np.where(df2['CurrentListing'].isnull(
        ) == False,  'CITES에 리스팅 정보가 존재합니다. 확인이 필요합니다.',  df2["Global_CITES_경고"])

        return (df2)


#     ### 중국 사이테스 분석
#     #Get Global Cites Word List

#     with open('/home/ec2-user/environment/efs/DATA/MATERIAL_SCREENING/01. DATA/chinese_cites_wordlist_dict.pkl', 'rb') as f:
#         chines_cites_wordlist_file = pickle.load(f)

#     # Get Global Cites File

#     chinese_cites = pd.read_excel('/home/ec2-user/environment/efs/DATA/MATERIAL_SCREENING/01. DATA/chinese_cites_file.xlsx')

    def Checking_Chinese_CITES(df, chines_cites_wordlist_file, chinese_cites):
        # 변수 안에 있는 단어의 리스트가 맞는지 체크 하는것 있는것만 가져와라는 형식입니다.
        pat = '|'.join(r"\b{}\b".format(x) for x in chines_cites_wordlist_file)

        # 매칭시 있을 수 있는 단어들의 대소문자 차이를 없애기 위해 모두 소문자로 처리
        df['cites_match_word_cn'] = df["영문명"].str.lower(
        ).str.findall(pat).apply(" ".join)

        # 변수 생성
        df2 = pd.merge(df, chinese_cites, on='cites_match_word_cn', how='left')

        # 경고문구 제작
        df2["CN_CITES_경고"] = ''
        df2["CN_CITES_경고"] = np.where(
            df2['cites_match_word_cn'] != "",  'CITES 확인이 필요합니다',  df2["CN_CITES_경고"])
        df2["CN_CITES_경고"] = np.where(df2['Chinese_cites_name'].isnull(
        ) == False,  'CITES에 리스팅 정보가 존재합니다. 확인이 필요합니다.',  df2["CN_CITES_경고"])

        return (df2)

    def Wash_perc_non_wash_perc(df):
        df['actual_wt_valid'] = pd.to_numeric(
            df['actual_wt_valid'], errors='coerce')
        df['wash_max_perc2'] = pd.to_numeric(
            df['wash_max_perc'], errors='coerce')
        df['non_wash_max_perc2'] = pd.to_numeric(
            df['non_wash_max_perc'], errors='coerce')
        df['actual_wt_valid_gr'] = df.groupby(
            ['성분명'])['actual_wt_valid'].transform(sum)
        df['비고'] = df['비고'].fillna("")
        df['Chinese_name'] = df['Chinese_name'].fillna("")
        df['wash_max_perc'] = df['wash_max_perc'].fillna("")
        df['non_wash_max_perc'] = df['non_wash_max_perc'].fillna("")

        # / 로 표기된건 비고를 쓰도록 하고 나머지만 비교 해야함
        df['wash_off_최대사용치_초과여부'] = ""
        df['wash_off_최대사용치_초과여부'] = np.where(
            df['wash_max_perc'] == '/', df['비고'],  df['wash_off_최대사용치_초과여부'])
        df['wash_off_최대사용치_초과여부'] = np.where((df['wash_off_최대사용치_초과여부'] == '') & (
            df['비고'] != ""), df['비고'], df['wash_off_최대사용치_초과여부'])
        df['wash_off_최대사용치_초과여부'] = np.where((df['wash_off_최대사용치_초과여부'] == '') & (
            df['wash_max_perc'] == "") & (df['Chinese_name'] != ""), '자료없음',  df['wash_off_최대사용치_초과여부'])
        df['wash_off_최대사용치_초과여부'] = np.where((df['wash_off_최대사용치_초과여부'] == '') & (df['actual_wt_valid_gr'] > df['wash_max_perc2']) & (
            df['wash_max_perc2'].isnull() == False) & (df['actual_wt_valid_gr'].isnull() == False), '초과',  df['wash_off_최대사용치_초과여부'])
        df['wash_off_최대사용치_초과여부'] = np.where((df['wash_off_최대사용치_초과여부'] == '') & (
            df['Chinese_name'] == ""), '확인필요', df['wash_off_최대사용치_초과여부'])

        df['leave_on_최대사용치_초과여부'] = ""
        df['leave_on_최대사용치_초과여부'] = np.where(
            df['non_wash_max_perc'] == '/', df['비고'],  df['leave_on_최대사용치_초과여부'])
        df['leave_on_최대사용치_초과여부'] = np.where((df['leave_on_최대사용치_초과여부'] == '') & (
            df['비고'] != ""), df['비고'], df['leave_on_최대사용치_초과여부'])
        df['leave_on_최대사용치_초과여부'] = np.where((df['leave_on_최대사용치_초과여부'] == '') & (
            df['non_wash_max_perc'] == "") & (df['Chinese_name'] != ""), '자료없음',  df['leave_on_최대사용치_초과여부'])
        df['leave_on_최대사용치_초과여부'] = np.where((df['leave_on_최대사용치_초과여부'] == '') & (df['actual_wt_valid_gr'] > df['non_wash_max_perc2']) & (
            df['non_wash_max_perc2'].isnull() == False) & (df['actual_wt_valid_gr'].isnull() == False), '초과',  df['leave_on_최대사용치_초과여부'])
        df['leave_on_최대사용치_초과여부'] = np.where((df['leave_on_최대사용치_초과여부'] == '') & (
            df['Chinese_name'] == ""), '확인필요', df['leave_on_최대사용치_초과여부'])

        return (df)

    # 1차 기술규범 판단 가장초기 데이터를 넣는다(중국에서 사용가능한지 여부 체크) 2차 테스트 완료 자료를 넣으면 됩니다.
    # run function
    second_fin_file_1 = Not_use_ingred_china(df)

    # 2차 사용제한 물질인지 체크합니다. 표3의 자료입니다
    third_file_1_ = third_table3_stage(second_fin_file_1, rule_table3_limit_)

    # 3차 방부제를 테스트 합니다. 표 4의 자료입니다.
    third_file_2_ = third_table4_stage(third_file_1_, table_4_)

    # 4차 자외선 차단제를 테스트 합니다. 표 5의 자료입니다.
    third_file_3_ = third_table5_stage(third_file_2_, table_5_)

    # 5차 준용착색제를 매칭합니다.
    third_file_4_1_ = Get_CI_name_data(
        third_file_3_, cosing_annexfile, ci_datas2)

    # 6차 준용착색제를 테스트 합니다. 이건 표6의 자료입니다.    s
    third_file_4_2_ = third_table6_stage(third_file_4_1_, table_6_)

    # 7차 사용목적 합하기
    third_file_4_2_ = Find_Issue_U_Purpose(third_file_4_2_)

    # 7-1차 사용목적 체크 및 합하고 이슈 제기
    third_file_5_1_ = Ingredient_function_purpose_Check_(
        third_file_4_2_, purpose_list)

    # 8차 사용가능한 특수 효능 찾아보기
    third_file_5_2_ = Find_Usable_effect(third_file_5_1_)

    # 9차 CI코드 있으면 바꾸기
    third_file_5_3_ = Change_CI_code_bfunction(third_file_5_2_)

    # 10차 목적에 따른 이슈 업데이트
    third_file_5_4_ = Change_purpose_issue_update(third_file_5_3_)

    # 11차 글로벌 사이테스 테스트
    third_file_6_ = Checking_Global_CITES(
        third_file_5_4_, cites_wordlist_file, g_cites_data)

    # 12차 중국 사이테스 테스트
    third_file_7_ = Checking_Chinese_CITES(
        third_file_6_, chines_cites_wordlist_file, chinese_cites)

    # 마지막으로 빠트린 wash off 등 테스트

    third_file_8_ = Wash_perc_non_wash_perc(third_file_7_)

    return (third_file_8_)

# 7.


ewgdata_g = call_csv(
    bucket_nm=BUCKET, key_nm="ewg-skindeep/fullfile/ewg_full_file.csv")


def Last_merge_file(thirdfile, ewgdata_g, hwahae_meiri2_cd, status):
    def Standard_name_error_code(thirdfile):
        thirdfile['표준화 명칭 오류'] = ""
        # 먼저 국문이랑 성분명 다르면 성분명이 나오게 끔 해야함
        thirdfile['표준화 명칭 오류'] = np.where(
            thirdfile['성분명'] != thirdfile['KoreanName'], thirdfile['성분명'], thirdfile['표준화 명칭 오류'])
        # 먼저 국문에 맞는 성분명이 없으면 없다고 나오게 끔 해야함
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['KoreanName'].isnull() == True) | (
            thirdfile['KoreanName'] == ""), "매칭되는 국문명칭없음", thirdfile['표준화 명칭 오류'])
        # 다음으로 중문명이랑 다르고 변경된 중문명이 비어있으면, 오류가 떴다고 알려줘야함
        # 초기변수들 nan value setting
        thirdfile['Chinese_name'] = thirdfile['Chinese_name'].fillna("")
        thirdfile['ChineseName'] = thirdfile['ChineseName'].fillna("")
        thirdfile['INCI'] = thirdfile['INCI'].fillna("")

        # 중문명칭은 나왔는데 실제 중문명칭이랑 다를 경우
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['ChineseName'].str.strip() != thirdfile['Chinese_name'].str.strip()) & (thirdfile['Chinese_name'] != "") & (
            thirdfile['표준화 명칭 오류'] != ""), thirdfile['표준화 명칭 오류'] + str("/") + thirdfile['Chinese_name'], thirdfile['표준화 명칭 오류'])
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['ChineseName'].str.strip() != thirdfile['Chinese_name'].str.strip()) & (
            thirdfile['Chinese_name'] != "") & (thirdfile['표준화 명칭 오류'] == ""), thirdfile['Chinese_name'], thirdfile['표준화 명칭 오류'])

        # 중문명칭은 나왔는데 중문명이 없는경우
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['ChineseName'].str.strip() != thirdfile['Chinese_name'].str.strip()) & (
            thirdfile['Chinese_name'] == "") & (thirdfile['표준화 명칭 오류'] != ""), thirdfile['표준화 명칭 오류'] + str("/ 매칭되는 제품중문명칭없음"), thirdfile['표준화 명칭 오류'])
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['ChineseName'].str.strip() != thirdfile['Chinese_name'].str.strip()) & (
            thirdfile['Chinese_name'] == "") & (thirdfile['표준화 명칭 오류'] == ""), str("매칭되는 제품중문명칭없음"), thirdfile['표준화 명칭 오류'])

        # 중문명칭도 없는 경우
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['ChineseName'].str.strip() == "") & (thirdfile['Chinese_name'] == "") & (
            thirdfile['표준화 명칭 오류'] != ""), thirdfile['표준화 명칭 오류'] + str("/ 매칭되는 제품중문명칭없음"), thirdfile['표준화 명칭 오류'])
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['ChineseName'].str.strip() == "") & (
            thirdfile['Chinese_name'] == "") & (thirdfile['표준화 명칭 오류'] == ""), str("매칭되는 제품중문명칭없음"), thirdfile['표준화 명칭 오류'])

        # INCI Name 차이 확인
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['INCI'].str.strip().str.lower() != thirdfile['영문명'].str.strip().str.lower()) & (
            thirdfile['영문명'] != "") & (thirdfile['표준화 명칭 오류'] != ""), thirdfile['표준화 명칭 오류'] + str("/ ") + thirdfile['영문명'], thirdfile['표준화 명칭 오류'])
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['INCI'].str.strip().str.lower() != thirdfile['영문명'].str.strip(
        ).str.lower()) & (thirdfile['영문명'] != "") & (thirdfile['표준화 명칭 오류'] == ""), thirdfile['영문명'], thirdfile['표준화 명칭 오류'])

        # INCI Name 차이 확인 중 비어있으면?
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['INCI'].str.strip().str.lower() != thirdfile['영문명'].str.strip().str.lower()) & (
            thirdfile['영문명'] == "") & (thirdfile['표준화 명칭 오류'] != ""), thirdfile['표준화 명칭 오류'] + str("/ 매칭되는 제품영문명칭없음"), thirdfile['표준화 명칭 오류'])
        thirdfile['표준화 명칭 오류'] = np.where((thirdfile['INCI'].str.strip().str.lower() != thirdfile['영문명'].str.strip().str.lower()) & (
            thirdfile['영문명'] == "") & (thirdfile['표준화 명칭 오류'] == ""), str("매칭되는 제품영문명칭없음"), thirdfile['표준화 명칭 오류'])

        return (thirdfile)

    def CAS_NO_Warning(thirdfile2):
        # 띄어쓰기 오류가 있는 경우가 있음
        thirdfile2['CAS No2'] = thirdfile2['CAS No'].str.replace(
            r' \(', r'(', regex=True)
        thirdfile2['CAS_list'] = thirdfile2['CAS No2'].str.split(" ")

        def CAS_list_check(df2, x):
            try:
                df2['CAS_No'] = df2['CAS_No'].str.strip()
            except:
                df2['CAS_No'] = ""
            c_no = df2['CAS_No'][x]
            c_list = df2['CAS_list'][x]
            try:
                if c_no in c_list:
                    if len(c_list) > 1:
                        c_list.remove(c_no)
                        joined_str = ", ".join(c_list)
                        joined_str = str("[복수존재] ") + joined_str
                        return joined_str
                    else:
                        return ''
                else:
                    if len(c_list) > 1:
                        joined_str = ", ".join(c_list)
                        joined_str = str("[정정, 복수존재] ") + joined_str
                        return joined_str
                    else:
                        return f'[정정] {c_list[0]}'

            except TypeError:
                return "확인불가"

        cas_nos = []
        for i in range(0, len(thirdfile2)):
            thirdfile2_1 = thirdfile2.copy()
            cas_nos_i = CAS_list_check(thirdfile2_1, i)
            cas_nos.append(cas_nos_i)

        thirdfile2['CAS NO'] = cas_nos
        thirdfile2['CAS_list'] = thirdfile2['CAS No2'].str.split(" ")
        thirdfile2['성분명'] = thirdfile2['성분명'].fillna("")
        thirdfile2['CAS NO'] = np.where((thirdfile2['CAS NO'] == "확인불가") & (
            thirdfile2['성분명'] != ""), "CAS NO 없음", thirdfile2['CAS NO'])
        thirdfile2['CAS NO'] = np.where((thirdfile2['CAS NO'] == "CAS NO 없음") & (
            thirdfile2['성분명'] == "정보없음"), "확인불가", thirdfile2['CAS NO'])

        return (thirdfile2)

    def sums_Name_change(df):
        df2 = df.copy()
        renames_col = {'material_sum': '단일성분표 기준 함량 합계', 'material_in_sum': '복합원료 성분비 함량 합계', 'actual_wt_valid_check': '복합성분표 기준 함량 합계', 'wash_off_최대사용치_초과여부': '최고역사량 초과 (wash off)',
                       'leave_on_최대사용치_초과여부': '최고 역사량 초과 (leave on)', '방부제챌린지테스트': '챌린지 테스트 필요 여부'}

        df2 = df2.rename(columns=renames_col)

        df2['단일성분표 기준 함량 합계'] = np.where(
            df2['단일성분표 기준 함량 합계'] == 'check', "", "100% 아님")
        df2['복합원료 성분비 함량 합계'] = np.where(
            df2['복합원료 성분비 함량 합계'] == 'check', "", "100% 아님")
        df2['복합성분표 기준 함량 합계'] = np.where(
            df2['복합성분표 기준 함량 합계'] == 'Check', "", df2["actual_wt_valid"])

        return df2

    def wash_leave_tag(df):
        thirdfile3 = df.copy()
        thirdfile3['color_index'] = thirdfile3['color_index'].fillna("")
        thirdfile3['최고역사량 초과 (wash off)'] = thirdfile3['최고역사량 초과 (wash off)'].fillna(
            "")
        thirdfile3['최고 역사량 초과 (leave on)'] = thirdfile3['최고 역사량 초과 (leave on)'].fillna(
            "")

        thirdfile3['최고역사량 초과 (wash off)'] = np.where((thirdfile3['최고역사량 초과 (wash off)'].str.contains(
            '화장품안전기술규범')), '화장품안전기술규범 참고', thirdfile3['최고역사량 초과 (wash off)'])
        thirdfile3['최고역사량 초과 (wash off)'] = np.where((thirdfile3['최고역사량 초과 (wash off)'] == '자료없음') & (
            thirdfile3['color_index'] != ""), '화장품안전기술규범 참고', thirdfile3['최고역사량 초과 (wash off)'])
        thirdfile3['최고역사량 초과 (wash off)'] = np.where((thirdfile3['최고역사량 초과 (wash off)'] == '확인필요') & (
            thirdfile3['color_index'] != ""), '화장품안전기술규범 참고', thirdfile3['최고역사량 초과 (wash off)'])

        thirdfile3['최고 역사량 초과 (leave on)'] = np.where((thirdfile3['최고 역사량 초과 (leave on)'].str.contains(
            '화장품안전기술규범')), '화장품안전기술규범 참고', thirdfile3['최고 역사량 초과 (leave on)'])
        thirdfile3['최고 역사량 초과 (leave on)'] = np.where((thirdfile3['최고 역사량 초과 (leave on)'] == '자료없음') & (
            thirdfile3['color_index'] != ""), '화장품안전기술규범 참고', thirdfile3['최고 역사량 초과 (leave on)'])
        thirdfile3['최고 역사량 초과 (leave on)'] = np.where((thirdfile3['최고 역사량 초과 (leave on)'] == '확인필요') & (
            thirdfile3['color_index'] != ""), '화장품안전기술규범 참고', thirdfile3['최고 역사량 초과 (leave on)'])

        return (thirdfile3)

    def Safety_rules_check(df):

        df['화장품안전기술규범대상'] = ''
        # 방부제 체크
        df['표4화장품준용방부제여부'] = df['표4화장품준용방부제여부'].fillna('')
        df['화장품안전기술규범대상'] = np.where(
            (df['표4화장품준용방부제여부'] != ""), '중국준용방부제', df['화장품안전기술규범대상'])

        # 준용 자외선 여부 체크
        df['표5화장품준용자외선여부'] = df['표5화장품준용자외선여부'].fillna('')
        df['화장품안전기술규범대상'] = np.where((df['표5화장품준용자외선여부'] != "") & (
            df['화장품안전기술규범대상'] != ''),  df['화장품안전기술규범대상'] + str('/ 중국준용자외선차단제'), df['화장품안전기술규범대상'])
        df['화장품안전기술규범대상'] = np.where((df['표5화장품준용자외선여부'] != "") & (
            df['화장품안전기술규범대상'] == ''),   str('중국준용자외선차단제'), df['화장품안전기술규범대상'])

        # 준용 자외선 여부 체크
        df['화장품안전기술규범대상'] = np.where((df['준용착색제 존재여부'] == 1) & (
            df['화장품안전기술규범대상'] != ''),   df['화장품안전기술규범대상'] + str('/ 중국준용착색제'), df['화장품안전기술규범대상'])
        df['화장품안전기술규범대상'] = np.where((df['준용착색제 존재여부'] == 1) & (
            df['화장품안전기술규범대상'] == ''),   str('중국준용착색제'), df['화장품안전기술규범대상'])

        return (df)

    def unique_purpose_rename(df):
        df['복합원료=1개의 원료사용목적'] = ''
        df['성분목적 이슈'] = df['성분목적 이슈'].fillna('')
        df['복합원료=1개의 원료사용목적'] = np.where(df['성분목적 이슈'] != "", "원료사용목적 정정", '')
        return (df)

    def Purpose_Error_check(status, df):

        df['원료사용목적 오류'] = ''
        df['원료사용목적 오류'] = np.where(
            df['Ingredient_Function lw'] == "", '검토불가', df['원료사용목적 오류'])
        df['원료사용목적 오류'] = np.where((df['대응 영문 Function'] == "") & (
            df['원료사용목적 오류'] != "검토불가"), '대응되는 영문사용목적 없음', df['원료사용목적 오류'])

        if status['특수/비특수'][0] == '특수':
            print("특수제품")
            df['원료사용목적 오류'] = np.where(
                df['Ingredient_Function lw'] == 'skin-conditioning agent', '특수제품사용불가', df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where(
                df['Ingredient_Function lw'] == 'skin conditioning agent', '특수제품사용불가',  df['원료사용목적 오류'])
        elif status['특수/비특수'][0] == '':
            df['원료사용목적 오류'] = df['원료사용목적 오류'] + str('특수제품 검토불가')
        else:
            print("비특수제품")

        # 특수제품 사용불가 여부 확인

        # 유화제 사용여부 확인해서 유화제인경우 두개의 목적 사용체크 하도록
        if status['유화제여부'][0] == False:
            # 유화제품이 아닌경우 유화제 사용 불가
            print('유화제 아님')
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying agent') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'),  df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying agent') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying stabilizer') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'), df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying stabilizer') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion stabilizer') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'), df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion stabilizer') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion agent') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'), df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion agent') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

        elif status['유화제여부'][0] == "False":
            # 유화제품이 아닌경우 유화제 사용 불가
            print('유화제 아님')
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying agent') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'),  df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying agent') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying stabilizer') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'), df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsifying stabilizer') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion stabilizer') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'), df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion stabilizer') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion agent') & (
                df['원료사용목적 오류'] != ''), df['원료사용목적 오류'] + str('/ 유화제품 外 사용불가'), df['원료사용목적 오류'])
            df['원료사용목적 오류'] = np.where((df['Ingredient_Function lw'] == 'emulsion agent') & (
                df['원료사용목적 오류'] == ''), '유화제품 外 사용불가',  df['원료사용목적 오류'])

        elif status['유화제여부'][0] == '':
            df['원료사용목적 오류'] = df['원료사용목적 오류'] + str('유화제 검토불가')
        else:
            print(status['유화제여부'][0])
            print('유화제임')

        # 대응되는 영문사용목적 체크
        hyo = status['제품효능'][0]
        # 미백 체크
        print(f'제품의 효능 : {hyo}입니다')
        df['효능관련 원료사용목적 오류'] = ""
        if status['제품효능'][0] == "미백":
            if 'Whitening Cosmetics' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            elif 'Whitening Agent' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            else:
                df['효능관련 원료사용목적 오류'] = '원료사용목적에 whitening agent  없음'

        # 자외선 체크
        elif status['제품효능'][0] == "자외선차단":
            if 'Sunscreen Agent' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            else:
                df['효능관련 원료사용목적 오류'] = '원료사용목적에 Sunscreen Agent 없음'

        # 청결(클렌저)
        elif status['제품효능'][0] == "청결(클렌저)":
            if 'Cleansing Agent' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            else:
                df['효능관련 원료사용목적 오류'] = '원료사용목적에 cleansing agent 없음'

        # 각질제거 체크
        elif status['제품효능'][0] == "각질제거":
            if 'Exfoliant' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            elif 'Abrasive' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            else:
                df['효능관련 원료사용목적 오류'] = '원료사용목적에 exfoliant 혹은 abrasive 없음'

        # 피막형성 체크
        elif status['제품효능'][0] == "픽서":
            if 'Film Forming' in df['대응 영문 Function'].unique():
                df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
            else:
                df['효능관련 원료사용목적 오류'] = '원료사용목적에  flim forming (agent)  없음'

        # 미백 및 자외선차단  체크
        elif status['제품효능'][0] == "미백 및 자외선차단":
            if 'Whitening Cosmetics' in df['대응 영문 Function'].unique():
                if 'Sunscreen Agent' in df['대응 영문 Function'].unique():
                    df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
                else:
                    df['효능관련 원료사용목적 오류'] = '원료사용목적에 Sunscreen Agent 없음'
            elif 'Whitening Agent' in df['대응 영문 Function'].unique():
                if 'Sunscreen Agent' in df['대응 영문 Function'].unique():
                    df['효능관련 원료사용목적 오류'] = df['효능관련 원료사용목적 오류']
                else:
                    df['효능관련 원료사용목적 오류'] = '원료사용목적에 Sunscreen Agent 없음'
            else:
                df['효능관련 원료사용목적 오류'] = '원료사용목적에 whitening agent 및  Sunscreen Agent 없음'

        elif status['제품효능'][0] == '':
            df['효능관련 원료사용목적 오류'] = ''

        else:
            df['효능관련 원료사용목적 오류'] = ''

        return (df)

    def purpose_rulebook_check(dft):
        df = dft.copy()
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = ""

        # 방부제 체크 있으면
        df['표4화장품준용방부제여부'] = df['표4화장품준용방부제여부'].fillna("")
        df['표4화장품준용방부제여부_n'] = 0
        df.loc[(df['표4화장품준용방부제여부']) != "", '표4화장품준용방부제여부_n'] = 1

        # 순번별 체크
        df['표4화장품준용방부제여부_농도초과여부'] = df['표4화장품준용방부제여부_농도초과여부'].fillna("")
        df['표4코드_그룹별최대값'] = df.groupby(['순번'])['표4화장품준용방부제여부_n'].transform(max)
        df.loc[(df['대응 영문 Function'] == "Preservative") & (
            df['표4코드_그룹별최대값'] == 0), '화장품안전기술규범 위반 원료 사용 목적 오류'] = '방부제 사용 불가'
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Preservative") & (df['표4코드_그룹별최대값'] > 0) & (df['표4화장품준용방부제여부_농도초과여부'] != "") & (
            df['화장품안전기술규범 위반 원료 사용 목적 오류'] != ""), df['화장품안전기술규범 위반 원료 사용 목적 오류'] + str("/ 방부제 농도 초과"), df['화장품안전기술규범 위반 원료 사용 목적 오류'])
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Preservative") & (df['표4코드_그룹별최대값'] > 0) & (
            df['표4화장품준용방부제여부_농도초과여부'] != "") & (df['화장품안전기술규범 위반 원료 사용 목적 오류'] == ""), str("방부제 농도 초과"), df['화장품안전기술규범 위반 원료 사용 목적 오류'])

        # 착색제 체크
        df['표6준용착색제여부_농도초과여부'] = df['표6준용착색제여부_농도초과여부'].fillna("")
        df['표6코드_그룹별최대값'] = df.groupby(['순번'])['준용착색제 존재여부'].transform(max)
        df.loc[(df['대응 영문 Function'] == "Colorant") & (
            df['표6코드_그룹별최대값'] == 0), '화장품안전기술규범 위반 원료 사용 목적 오류'] = '착색제 사용 불가'
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Colorant") & (df['표4코드_그룹별최대값'] > 0) & (df['표6준용착색제여부_농도초과여부'] != "") & (
            df['화장품안전기술규범 위반 원료 사용 목적 오류'] != ""), df['화장품안전기술규범 위반 원료 사용 목적 오류'] + str("/ ") + str("착색제 농도 초과"), df['화장품안전기술규범 위반 원료 사용 목적 오류'])
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Colorant") & (df['표4코드_그룹별최대값'] > 0) & (
            df['표6준용착색제여부_농도초과여부'] != "") & (df['화장품안전기술규범 위반 원료 사용 목적 오류'] == ""), str("착색제 농도 초과"), df['화장품안전기술규범 위반 원료 사용 목적 오류'])
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Colorant") & (df['준용착색제 존재여부'] > 0) & (
            df['화장품안전기술규범 위반 원료 사용 목적 오류'] != ""), df['화장품안전기술규범 위반 원료 사용 목적 오류'] + str("/ ") + df['CI_name'], df['화장품안전기술규범 위반 원료 사용 목적 오류'])
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Colorant") & (df['준용착색제 존재여부'] > 0) & (
            df['화장품안전기술규범 위반 원료 사용 목적 오류'] == ""), df['CI_name'], df['화장품안전기술규범 위반 원료 사용 목적 오류'])

        # 자외선차단제 원료 확인
        df['표5화장품준용자외선여부_농도초과여부'] = df['표5화장품준용자외선여부_농도초과여부'].fillna("")
        df['표5화장품준용자외선여부'] = df['표5화장품준용자외선여부'].fillna("")
        df['표5화장품준용자외선여부_n'] = 0
        df.loc[(df['표5화장품준용자외선여부_n']) != "", '표5화장품준용자외선여부_n'] = 1

        df['표5코드_그룹별최대값'] = df.groupby(['순번'])['표5화장품준용자외선여부_n'].transform(max)

        df.loc[(df['대응 영문 Function'] == "Sunscreen Agent") & (
            df['표5코드_그룹별최대값'] == 0), '화장품안전기술규범 위반 원료 사용 목적 오류'] = '자외선차단제 사용 불가'
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Sunscreen Agent") & (df['표5코드_그룹별최대값'] > 0) & (df['표5화장품준용자외선여부_농도초과여부'] != "") & (
            df['화장품안전기술규범 위반 원료 사용 목적 오류'] != ""), df['화장품안전기술규범 위반 원료 사용 목적 오류'] + str("/ ") + str("자외선차단제 농도 초과"), df['화장품안전기술규범 위반 원료 사용 목적 오류'])
        df['화장품안전기술규범 위반 원료 사용 목적 오류'] = np.where((df['대응 영문 Function'] == "Sunscreen Agent") & (df['표5코드_그룹별최대값'] > 0) & (
            df['표5화장품준용자외선여부_농도초과여부'] != "") & (df['화장품안전기술규범 위반 원료 사용 목적 오류'] == ""), str("자외선차단제 농도 초과"), df['화장품안전기술규범 위반 원료 사용 목적 오류'])

        return (df)

    def CITES_rename(df):
        df = df.rename(columns={'CN_CITES_경고': '중국위기동식물 경고', 'Global_CITES_경고': 'Global CITES 경고',
                       'cites_match_word_cn': '중국위기동식물경고 의심단어', 'cites_match_word': 'Global CITES 경고 의심단어'})
        return df

    def Make_EWG_last_file(thirdfile3_e5, ewgdata_g):

        ewgdata_g['name_lw'] = ewgdata_g['name'].str.strip(
        ).str.lower().str.replace(" ", "").str.replace("-", "")
        thirdfile3_e5['영문명_lw'] = thirdfile3_e5['영문명'].str.strip(
        ).str.lower().str.replace(" ", "").str.replace("-", "")

        thirdfile4 = pd.merge(thirdfile3_e5, ewgdata_g[[
                              'name_lw', 'EWG']], left_on='영문명_lw', right_on='name_lw', how='left')

        thirdfile4.loc[thirdfile4['EWG'].isnull() == True,
                       'EWG'] = "EWG 데이터 확인 불가"

        return (thirdfile4)

    def Extract_Finder(df):
        df['추출물 부위'] = ""
        df.loc[(df['성분명'].str.contains('추출') == True) & (df['영문명'].str.lower(
        ).str.contains('extract') == True),  '추출물 부위'] = '추출물 부위 기재 필요'
        df['추출물 부위'] = np.where((df['영문명'].str.lower().str.strip() == 'fragrance') & (
            df['추출물 부위'] != ""), df['추출물 부위'] + str('/IFRA Certificate'), df['추출물 부위'])
        df['추출물 부위'] = np.where((df['영문명'].str.lower().str.strip() == 'fragrance') & (
            df['추출물 부위'] == ""), str('IFRA Certificate'), df['추출물 부위'])
        return (df)

    import numpy as np

    def not_use_and_restrict(df):
        # 표 3에 포함되는지 여부
        df['표3화장품사용제한여부'] = df['표3화장품사용제한여부'].fillna("")
        df['중국사용불가원료'] = ""
        df['중국사용불가원료'] = np.where(
            df['표3화장품사용제한여부'] != "", "사용불가", df['중국사용불가원료'])
        # 화장품안전기술규범" 사용금지 성분' '"화장품안전기술규범" 사용금지 성분으로 조정 예정' 이 두가지 이 두가지에 포함되는지 체크
        df['중국사용불가원료'] = np.where(
            df['not_use_code'] == 1, "사용불가", df['중국사용불가원료'])

        return (df)

    thirdfile2 = Standard_name_error_code(thirdfile)

    thirdfile2 = CAS_NO_Warning(thirdfile2)

    thirdfile3 = sums_Name_change(thirdfile2)

    thirdfile3 = wash_leave_tag(thirdfile3)

    thirdfile3_d = Safety_rules_check(thirdfile3)

    thirdfile3_e = unique_purpose_rename(thirdfile3_d)

    thirdfile3_e3 = Purpose_Error_check(status, thirdfile3_e)

    thirdfile3_e4 = purpose_rulebook_check(thirdfile3_e3)

    thirdfile3_e5 = CITES_rename(thirdfile3_e4)

    thirdfile5 = Make_EWG_last_file(thirdfile3_e5, ewgdata_g)

    thirdfile6 = Extract_Finder(thirdfile5)

    thirdfile7 = not_use_and_restrict(thirdfile6)

    thirdfile8 = pd.merge(thirdfile7, hwahae_meiri2_cd,
                          left_on='성분코드', right_on='성분코드', how='left')
    # cir data
    cirdb = cir_safety[['mat_name', 'link']].rename(
        columns={'link': 'CIR_Link'})
    thirdfile8_1 = pd.merge(thirdfile8, cirdb, left_on=thirdfile8['영문명'].str.lower(
    ).str.strip(), right_on=cirdb['mat_name'].str.lower().str.strip(), how='left')

    return (thirdfile8_1)

# 8.


def Remain_cells(thirdfile6, status):
    thirdfile6['최고역사량 초과 (wash off)'] = thirdfile6['최고역사량 초과 (wash off)'].str.replace(
        'nan', '')
    thirdfile6['최고 역사량 초과 (leave on)'] = thirdfile6['최고 역사량 초과 (leave on)'].str.replace(
        'nan', '')
    thirdfile6['챌린지 테스트 필요 여부'] = thirdfile6['챌린지 테스트 필요 여부'].str.replace(
        'nan', '')
    if status['제품타입'][0] == 'leave_on':
        df = thirdfile6[["순번", "KoreanName", "ChineseName", "INCI", "rm_ing_fla", "ingre_in_rm", "actual_wt", "Ingredient_Function", "CAS_No", "Raw_t_name", "Raw_m_name", "COA", '표준화 명칭 오류', 'CAS NO',
                         '단일성분표 기준 함량 합계', '복합원료 성분비 함량 합계', '복합성분표 기준 함량 합계',  '중국사용불가원료',
                         '최고 역사량 초과 (leave on)', '챌린지 테스트 필요 여부', '화장품안전기술규범대상', '복합원료=1개의 원료사용목적', '원료사용목적 오류', '효능관련 원료사용목적 오류', '화장품안전기술규범 위반 원료 사용 목적 오류',
                         'Global CITES 경고',  'Global CITES 경고 의심단어', '중국위기동식물 경고', '중국위기동식물경고 의심단어',  'EWG', '화해', '메이리슈싱', '추출물 부위', 'CIR_Link']]

    elif status['제품타입'][0] == 'wash_off':
        df = thirdfile6[["순번", "KoreanName", "ChineseName", "INCI", "rm_ing_fla", "ingre_in_rm", "actual_wt", "Ingredient_Function", "CAS_No", "Raw_t_name", "Raw_m_name", "COA", '표준화 명칭 오류', 'CAS NO',
                         '단일성분표 기준 함량 합계', '복합원료 성분비 함량 합계', '복합성분표 기준 함량 합계',  '중국사용불가원료',   '최고역사량 초과 (wash off)',
                         '챌린지 테스트 필요 여부', '화장품안전기술규범대상', '복합원료=1개의 원료사용목적', '원료사용목적 오류', '효능관련 원료사용목적 오류', '화장품안전기술규범 위반 원료 사용 목적 오류',
                         'Global CITES 경고',  'Global CITES 경고 의심단어', '중국위기동식물 경고', '중국위기동식물경고 의심단어',  'EWG', '화해', '메이리슈싱', '추출물 부위', 'CIR_Link']]
    else:
        df = thirdfile6[["순번", "KoreanName", "ChineseName", "INCI", "rm_ing_fla", "ingre_in_rm", "actual_wt", "Ingredient_Function", "CAS_No", "Raw_t_name", "Raw_m_name", "COA", '표준화 명칭 오류', 'CAS NO',
                         '단일성분표 기준 함량 합계', '복합원료 성분비 함량 합계', '복합성분표 기준 함량 합계',  '중국사용불가원료',   '최고역사량 초과 (wash off)',
                         '최고 역사량 초과 (leave on)', '챌린지 테스트 필요 여부', '화장품안전기술규범대상', '복합원료=1개의 원료사용목적', '원료사용목적 오류', '효능관련 원료사용목적 오류', '화장품안전기술규범 위반 원료 사용 목적 오류',
                         'Global CITES 경고',  'Global CITES 경고 의심단어', '중국위기동식물 경고', '중국위기동식물경고 의심단어',  'EWG', '화해', '메이리슈싱', '추출물 부위', 'CIR_Link']]

    sub_dt = thirdfile6[(thirdfile6['최고 역사량 초과 (leave on)'] == "화장품안전기술규범 참고") | (thirdfile6['최고역사량 초과 (wash off)'] == "화장품안전기술규범 참고") | (
        thirdfile6['표3화장품사용제한여부'] != "") | (thirdfile6['표4화장품준용방부제여부'] != "") | (thirdfile6['표5화장품준용자외선여부'] != "") | (thirdfile6['준용착색제 존재여부'] != 0)]

    subs_ = sub_dt[['성분명', 'actual_wt_valid_gr', '적용 및 사용범위', '화장품 중 최대사용농도', '기타 제한조건 및 요구사항', '라벨에 반드시 표기해야하는 주의사항',
                    '주의사항_표3', '표3화장품사용제한_농도초과여부', '화장품 중 최대사용농도_방부제', '표4화장품준용방부제여부_농도초과여부', '사용범위및제한조건_방부제', '라벨에 반드시 표기해야하는 주의사항_방부제',
                    '주의사항_방부제',  '화장품 중 최대사용농도_자외선', '사용범위및제한조건_자외선', '라벨에 반드시 표기해야하는 주의사항_자외선', '주의사항_자외선',
                    '표5화장품준용자외선여부_농도초과여부', 'CI_Check', 'color_index', '표6준용착색제여부_농도초과여부', '기타제한조건 및 요구사항_착색제', '주의사항_착색제', '사용범위_착색제', 'color']]

    colrename = {'actual_wt_valid_gr': '원료사용비중', '적용 및 사용범위': '표3 사용범위', '화장품 중 최대사용농도': '표3 최대사용농도', '기타 제한조건 및 요구사항': '표3 제한조건', '라벨에 반드시 표기해야하는 주의사항': '표3 라벨표기', '주의사항_표3': '표3 주석',
                 '표3화장품사용제한_농도초과여부': '표3 농도초과여부', '화장품 중 최대사용농도_방부제': '표4 최대사용농도', '표4화장품준용방부제여부_농도초과여부': '표4 농도초과여부', '사용범위및제한조건_방부제': '표4 사용범위 및 제한조건', '라벨에 반드시 표기해야하는 주의사항_방부제': '표4 라벨표기',
                 '주의사항_방부제': '표4 주석', '화장품 중 최대사용농도_자외선': '표5 최대사용농도', '사용범위및제한조건_자외선': '표5 사용범위 및 제한조건', '라벨에 반드시 표기해야하는 주의사항_자외선': '표5 라벨표기', '주의사항_자외선': '표5 각주',
                 '표5화장품준용자외선여부_농도초과여부': '표5 농도초과여부', 'CI_Check': '표6 착색제코드 관련', 'color_index': '표6 착색제코드', '표6준용착색제여부_농도초과여부': '표6 농도초과여부', '기타제한조건 및 요구사항_착색제': '표6 제한조건 및 최대농도', '주의사항_착색제': '표6 주석',
                 '사용범위_착색제': '표6 사용범위', 'color': '표6 색상'}

    subs_2 = subs_.rename(columns=colrename)

    return (df, subs_2)

# 9.


def Make_output_Excel(fin_cell, sub_file, col_length, status):

    # border_style
    BORDER_NONE = None
    BORDER_DASHDOT = 'dashDot'
    BORDER_DASHDOTDOT = 'dashDotDot'
    BORDER_DASHED = 'dashed'
    BORDER_DOTTED = 'dotted'
    BORDER_DOUBLE = 'double'
    BORDER_HAIR = 'hair'
    BORDER_MEDIUM = 'medium'
    BORDER_MEDIUMDASHDOT = 'mediumDashDot'
    BORDER_MEDIUMDASHDOTDOT = 'mediumDashDotDot'
    BORDER_MEDIUMDASHED = 'mediumDashed'
    BORDER_SLANTDASHDOT = 'slantDashDot'
    BORDER_THICK = 'thick'
    BORDER_THIN = 'thin'

    borderbox = Border(left=Side(BORDER_THICK, color='FF000000'),
                       right=Side(BORDER_THICK, color='FF000000'),
                       top=Side(BORDER_THICK, color='FF000000'),
                       bottom=Side(BORDER_THICK, color='FF000000'))

    borderbox2 = Border(left=Side(BORDER_MEDIUM, color='FF000000'),
                        right=Side(BORDER_MEDIUM, color='FF000000'),
                        top=Side(BORDER_MEDIUM, color='FF000000'),
                        bottom=Side(BORDER_MEDIUM, color='FF000000'))
    borderbox3 = Border(left=Side(BORDER_THIN, color='FF868686'),
                        right=Side(BORDER_THIN, color='FF868686'),
                        top=Side(BORDER_THIN, color='FF868686'),
                        bottom=Side(BORDER_THIN, color='FF868686'))
    # a셀 너비 90픽셀

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "원료스크리닝 검토 결과"
    ws.sheet_properties.tabColor = 'E6B8B7'
    ws.insert_rows(1, 8)
    for r in dataframe_to_rows(fin_cell, index=False, header=True):
        ws.append(r)

    # 제품 스테이터스 기록
    ws['A2'] = '제품명'
    ws['B2'] = status['제품명'][0]
    ws['A3'] = '유화제여부'
    ws['B3'] = status['유화제여부'][0]
    ws['A4'] = '제품효능'
    ws['B4'] = status['제품효능'][0]
    ws['A5'] = '특수/비특수'
    ws['B5'] = status['특수/비특수'][0]
    ws['A6'] = '제품타입'
    ws['B6'] = status['제품타입'][0]

    # 스테이터스 셀 박스
    status_cells = ['A2', 'B2', 'A3', 'B3', 'A4', 'B4', 'A5', 'B5', 'A6', 'B6']
    for i in status_cells:
        ws[i].border = borderbox

    ###

    def merged_cell_border(row, start_col, last_col):
        # border check :
        for col_num in range(start_col, last_col+1):
            c2 = ws.cell(row=row, column=col_num)
            c2.border = borderbox2

    def merged_cell_border2(row, start_col, last_col):
        # border check :
        for col_num in range(start_col, last_col+1):
            c2 = ws_2.cell(row=row, column=col_num)
            c2.border = borderbox2

    # 위생허가 복합성분표 원본 정보
    # col_length 는 load 프로세스에서 가져옴

    ws.merge_cells(start_row=8, start_column=1,
                   end_row=9, end_column=col_length)
    c = ws.cell(row=8, column=1)
    c.value = '위생허가용 복합성분표 원본'
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = Font(b=True)
    greenFill = PatternFill(start_color="00EBF1DE",
                            end_color="00EBF1DE",
                            fill_type='solid')

    c.fill = greenFill
    merged_cell_border(row=8, start_col=1, last_col=col_length)
    merged_cell_border(row=9, start_col=1, last_col=col_length)

    for nrow in range(10, 11):
        for i in range(0, col_length):
            i2 = i + 1
            c = ws.cell(row=nrow, column=i2)
            c.font = Font(b=True)
            c.fill = greenFill
            c.border = borderbox2

    # 칼럼 아래 샐 (정보입력된 값들 ) 기초정보
    for i in range(0, col_length):
        for k in range(0, len(fin_cell)):
            t = i + 1
            k2 = k + 10 + 1
            c = ws.cell(row=k2, column=t)
            c.border = borderbox3

    # 여기서는 원료스크리닝 프로그램 결과 보여주기
    # fin_cell 은 최종 나온것
    ws.merge_cells(start_row=8, start_column=col_length + 1,
                   end_row=8, end_column=len(fin_cell.columns))
    c = ws.cell(row=8, column=col_length + 1)
    c.value = '원료 스크리닝 프로그램 검토 결과'
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = Font(b=True, color="00C00000")
    redFill = PatternFill(start_color='00FDE9D9',
                          end_color='00FDE9D9',
                          fill_type='solid')
    c.fill = redFill
    merged_cell_border(row=8, start_col=col_length + 1,
                       last_col=len(fin_cell.columns))

    # 기본정보 확인 2, 오류검증 3, 위생허가 원료스크리닝 9, 통관리스크 4, 마케팅 리스크 4
    if (status['제품타입'][0] == 'wash_off') | (status['제품타입'][0] == 'leave_on'):
        print(status['제품타입'][0])
        secon_merge = {"기본정보 확인": 2, "원료비중 오류검증": 3,
                       "위생허가 원료스크리닝": 8, "통관리스크": 4, "마케팅 리스크": 3, '추가정보': 2}
    else:
        secon_merge = {"기본정보 확인": 2, "원료비중 오류검증": 3,
                       "위생허가 원료스크리닝": 9, "통관리스크": 4, "마케팅 리스크": 3, '추가정보': 2}
    # 중간단위 머지
    e_col = col_length + 1
    for name in secon_merge:
        i = secon_merge[name]
        e_col = i + e_col
        ecol2 = e_col - 1
        s_col = e_col - i
        ws.merge_cells(start_row=9, start_column=s_col,
                       end_row=9, end_column=ecol2)
        c = ws.cell(row=9, column=s_col)
        c.value = name
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(b=True, color="00C00000")
        c.fill = redFill
        merged_cell_border(row=9, start_col=s_col, last_col=ecol2)

    # 마지막 칼럼들 색칠하기
    for i in range(col_length, len(fin_cell.columns)):
        t = i + 1
        c = ws.cell(row=10, column=t)
        c.font = Font(b=True, color="00C00000")
        c.fill = redFill
        c.border = borderbox2

    # 칼럼 아래 샐 (정보입력된 값들 )
    # 칼럼방향
    for i in range(col_length, len(fin_cell.columns)):
        # 로우 방향
        for k in range(0, len(fin_cell)):
            t = i + 1
            k2 = k + 10 + 1
            c = ws.cell(row=k2, column=t)
            c.font = Font(b=False, color="00C00000")
            c.border = borderbox3
    # row 수 세기
    num_count = fin_cell.groupby(['순번'])['KoreanName'].count().reset_index()
    num_count.rename(columns={'KoreanName': 'n_row'}, inplace=True)

    col_numb = 1
    for col in fin_cell.columns:
        if col == '챌린지 테스트 필요 여부':
            print('merge challenge test cells')
            st_row = 11
            ws.merge_cells(start_row=st_row, start_column=col_numb,
                           end_row=st_row+len(fin_cell)-1, end_column=col_numb)
            c = ws.cell(row=st_row, column=col_numb)
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif col == '효능관련 원료사용목적 오류':
            print('merge c효능관련 원료사용목적 오류 cells')
            st_row = 11
            ws.merge_cells(start_row=st_row, start_column=col_numb,
                           end_row=st_row+len(fin_cell)-1, end_column=col_numb)
            c = ws.cell(row=st_row, column=col_numb)
            c.alignment = Alignment(horizontal='center', vertical='center')

        elif col == '화장품안전기술규범 위반 원료 사용 목적 오류 체크':
            print('화장품안전기술규범 위반 원료 사용 목적 오류 체크 cells')
            st_row = 11
            for rn in range(0, len(num_count)):
                n_rows = num_count['n_row'][rn]
                if n_rows == 1:
                    c = ws.cell(row=st_row, column=col_numb)
                    c.alignment = Alignment(
                        horizontal='center', vertical='center')
                    st_row = st_row + 1
                else:
                    ws.merge_cells(start_row=st_row, start_column=col_numb,
                                   end_row=st_row+n_rows - 1, end_column=col_numb)
                    c = ws.cell(row=st_row, column=col_numb)
                    c.alignment = Alignment(
                        horizontal='center', vertical='center')
                    st_row = st_row + n_rows

        elif col == '단일성분표 기준 함량 합계':
            print('merge 단일성분표  cells')
            st_row = 11
            ws.merge_cells(start_row=st_row, start_column=col_numb,
                           end_row=st_row+len(fin_cell)-1, end_column=col_numb)
            c = ws.cell(row=st_row, column=col_numb)
            c.alignment = Alignment(horizontal='center', vertical='center')

        elif col == '복합원료 성분비 함량 합계':
            print('복합원료 cells')
            st_row = 11
            for rn in range(0, len(num_count)):
                n_rows = num_count['n_row'][rn]
                if n_rows == 1:
                    c = ws.cell(row=st_row, column=col_numb)
                    c.alignment = Alignment(
                        horizontal='center', vertical='center')
                    st_row = st_row + 1
                else:
                    ws.merge_cells(start_row=st_row, start_column=col_numb,
                                   end_row=st_row+n_rows - 1, end_column=col_numb)
                    c = ws.cell(row=st_row, column=col_numb)
                    c.alignment = Alignment(
                        horizontal='center', vertical='center')
                    st_row = st_row + n_rows

        elif col == '순번':
            print('순번 cells')
            st_row = 11
            for rn in range(0, len(num_count)):
                n_rows = num_count['n_row'][rn]
                if n_rows == 1:
                    c = ws.cell(row=st_row, column=col_numb)
                    c.alignment = Alignment(
                        horizontal='center', vertical='center')
                    st_row = st_row + 1
                else:
                    ws.merge_cells(start_row=st_row, start_column=col_numb,
                                   end_row=st_row+n_rows - 1, end_column=col_numb)
                    c = ws.cell(row=st_row, column=col_numb)
                    c.alignment = Alignment(
                        horizontal='center', vertical='center')
                    st_row = st_row + n_rows

        col_numb = col_numb + 1

    # 순번에 맞게 머지 하기

    ws_2 = wb.create_sheet("화장품안전기술규범 대상 별도표")
    ws_2.insert_rows(1, 9)
    ws_2.sheet_properties.tabColor = 'D8E4BC'
    for r in dataframe_to_rows(sub_file, index=False, header=True):
        ws_2.append(r)

    ws2_top_ = {'원료명 및 원료비중': 2,  '표3 사용제한': 6,
                '표4 방부제': 5, '표5 자외선 차단제': 5, '표6 준용착색제': 7}

    # 중간단위 머지
    e_col = 1
    for name in ws2_top_:
        i = ws2_top_[name]
        e_col = i + e_col
        ecol2 = e_col - 1
        s_col = e_col - i
        ws_2.merge_cells(start_row=10, start_column=s_col,
                         end_row=10, end_column=ecol2)
        c = ws_2.cell(row=10, column=s_col)
        c.value = name
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(b=True)
        c.fill = greenFill

    merged_cell_border2(row=10, start_col=1, last_col=len(sub_file.columns))
    merged_cell_border2(row=11, start_col=1, last_col=len(sub_file.columns))

    for i in range(0, len(sub_file.columns)):
        t = i + 1
        c = ws_2.cell(row=11, column=t)
        c.fill = greenFill

    # 칼럼 아래 샐 (정보입력된 값들 )
    # 칼럼방향
    for i in range(0, len(sub_file.columns)):
        # 로우 방향
        for k in range(0, len(sub_file)):
            t = i + 1
            k2 = k + 11 + 1
            c = ws_2.cell(row=k2, column=t)
            c.font = Font(b=False, color="00C00000")
            c.border = borderbox3

    ws_2['A8'] = '* 하단 표의 원료가 중국준용 방부제, 자외선차단제, 착색제로 사용되었을 경우에만 해당되며, 아닌 경우 무시해도 됩니다. 가령 티타늄디옥사이드가 착색제로 원료사용목적이 기제된 복합원료의 성분으로 처방된 경우에는 하단 표의 준용자외선 차단제 관련 내용은 적용되지 않습니다'
    ws_2['A9'] = '* 하단 표에 기재된 내용 外 해당원료의 화장품관리규범 관련 이슈는 없는 것으로 이해하면 됩니다.'

    return (wb)

# 10.


def Make_output_Info(fin_cell):
    """
    /**
    * 기능 : 화면 Display 용 데이터 추출
    * @param {fin_cell} 계산에 필요한 엑셀 데이터
    * @return {data} 결과 json 데이터

    */
    """
    result = {}
        # 카드 정보
    result["card-top-data"] = {}
    # 파일 정보
    result["file-info"] = {}
    # 카드 raw 정보
    result["prob-base-data"] = []
    result["prob-cpl-data"] = []
    result["prob-hyg-data"] = []
    result["prob-cc-data"] = []
    result["prob-mkt-data"] = []

    # 기본정보 에러
    baseErrCnt = 0

    # 기본정보 에러
    base_col0 = fin_cell.iloc[:, 0].name
    base_col1 = fin_cell.iloc[:, 1].name
    base_col2 = fin_cell.iloc[:, 12].name # 표준화 명칭 오류
    probBaseData = pd.DataFrame()
    for column_name, row in fin_cell[[base_col0, base_col1, base_col2]].iterrows():
        new_rows = []
        if (len(row[2]) == 0):
            new_rows = [[str(row[0]), row[1], 'success', 'success', 'success']]
        else:
            baseErrCnt += 1
            new_rows = [[str(row[0]), row[1], 'danger', 'danger', 'danger']]
        nr = pd.DataFrame(new_rows)
        probBaseData = pd.concat([probBaseData, nr], ignore_index=True)    
    result["card-top-data"][".base-err-cnt"] = str(baseErrCnt)
    result["prob-base-data"] = probBaseData.values.tolist()


    # 원료비중 오류
    baseMatErrCnt1 = 0
    baseMatErrCnt2 = 0
    baseMatErrCnt3 = 0
    
    # 원료비중 에러 raw data
    mat_col0 = fin_cell.iloc[:, 0].name
    mat_col1 = fin_cell.iloc[:, 1].name
    mat_col2 = fin_cell.iloc[:, 14].name # 단일성분표 기준 함량 합계
    mat_col3 = fin_cell.iloc[:, 15].name # 복합원료 성분비 함량 합계
    mat_col4 = fin_cell.iloc[:, 16].name # 복합성분표 기준 함량 합계
    probMatData = pd.DataFrame()
    for column_name, row in fin_cell[[mat_col0, mat_col1, mat_col2, mat_col3, mat_col4]].iterrows():
        cell1 = str(row[0])
        cell2 = row[1]
        cell3 = "danger"
        cell4 = "danger"
        cell5 = "danger"

        if (len(row[2]) == 0):
            cell3 = "success"
        else:
            baseMatErrCnt1 += 1

        if (len(row[3]) == 0):            
            cell4 = "success"
        else:
            baseMatErrCnt2 += 1

        if (len(row[4]) == 0):            
            cell5 = "success"
        else:
            baseMatErrCnt3 += 1

        new_rows = [[cell1, cell2, cell3, cell4, cell5]]
        nr = pd.DataFrame(new_rows)
        probMatData = pd.concat([probMatData, nr], ignore_index=True)        
    result["prob-cpl-data"] = probMatData.values.tolist()
    baseMatErrCntSum = sum([baseMatErrCnt1, baseMatErrCnt2, baseMatErrCnt3])
    result["card-top-data"][".cpl-err"] = "문제 있음" if baseMatErrCntSum > 0 else "문제 없음"

    # 위생허가 구제 오류
    hygieneErrCnt1 = 0
    hygieneErrCnt2 = 0
    hygieneErrCnt3 = 0
    hygieneErrCnt4 = 0
    hygieneErrCnt5 = 0
    hygieneErrCnt6 = 0
    hygieneErrCnt7 = 0
    hygieneErrCnt8 = 0
    hygieneErrCnt9 = 0

    # 위생허가 구제 오류 raw data
    hyg_col0 = fin_cell.iloc[:, 0].name
    hyg_col1 = fin_cell.iloc[:, 1].name
    hyg_col2 = fin_cell.iloc[:, 17].name    # 중국사용불가원료
    hyg_col3 = fin_cell.iloc[:, 18].name    # 최고역사량 초과 (wash off)
    hyg_col4 = fin_cell.iloc[:, 19].name    # 최고 역사량 초과 (leave on)
    hyg_col5 = fin_cell.iloc[:, 20].name    # 챌린지 테스트 필요 여부
    hyg_col6 = fin_cell.iloc[:, 21].name    # 화장품안전기술규범대상
    hyg_col7 = fin_cell.iloc[:, 22].name    # 복합원료=1개의 원료사용목적
    hyg_col8 = fin_cell.iloc[:, 23].name    # 원료사용목적 오류
    hyg_col9 = fin_cell.iloc[:, 24].name    # 효능관련 원료사용목적 오류
    hyg_col10 = fin_cell.iloc[:, 25].name   # 화장품안전기술규범 위반 원료 사용 목적 오류
    probHygData = pd.DataFrame()
    for column_name, row in fin_cell[[hyg_col0, hyg_col1, hyg_col2, hyg_col3, hyg_col4, hyg_col5, hyg_col6, hyg_col7, hyg_col8, hyg_col9, hyg_col10]].iterrows():
        # print(str(row[2]) + " " + str(row[3]) + " " + str(row[4]) + " " + str(row[5]))
        cell1 = str(row[0])
        cell2 = row[1]
        cell3 = "danger"
        cell4 = "danger"
        cell5 = "danger"
        
        if (len(row[3]) == 0 and len(row[4]) == 0):
            cell4 = "success"

        if (len(row[2]) == 0):
            cell3 = "success"
        else:
            hygieneErrCnt1 += 1

        if (len(row[3]) > 0):
            hygieneErrCnt2 += 1

        if (len(row[4]) > 0):
            hygieneErrCnt3 += 1

        if (len(row[5]) > 0):
            hygieneErrCnt4 += 1

        if (len(row[6]) > 0):
            hygieneErrCnt5 += 1

        if (len(row[7]) > 0):
            hygieneErrCnt6 += 1

        if (len(row[8]) > 0):
            hygieneErrCnt7 += 1

        if (len(row[9]) > 0):
            hygieneErrCnt8 += 1

        if (len(row[10]) == 0):
            cell5 = "success"
        else:
            hygieneErrCnt9 += 1

        new_rows = [[cell1, cell2, cell3, cell4, cell5]]
        nr = pd.DataFrame(new_rows)
        probHygData = pd.concat([probHygData, nr], ignore_index=True)        
    result["prob-hyg-data"] = probHygData.values.tolist()
    hygieneErrCntSum = sum([hygieneErrCnt1, hygieneErrCnt2, hygieneErrCnt3, hygieneErrCnt4, hygieneErrCnt5, hygieneErrCnt6, hygieneErrCnt7, hygieneErrCnt8, hygieneErrCnt9])
    if (hygieneErrCnt4 > 0) :
        result["card-top-data"][".challenge"] = "검토 필요"
    else:
        result["card-top-data"][".challenge"] = "검토 필요 없음"
    if (hygieneErrCnt7 > 0):
        result["card-top-data"][".materials"] = "검토 필요"
    else:
        result["card-top-data"][".materials"] = "검토 필요 없음"

    result["card-top-data"][".hyg-err"] = "문제 있음" if hygieneErrCntSum > 0 else "문제 없음"

    # 통관 위험
    ccErrCnt = 0

    # 통관 위험 raw data
    cc_col0 = fin_cell.iloc[:, 0].name
    cc_col1 = fin_cell.iloc[:, 1].name
    cc_col2 = fin_cell.iloc[:, 27].name # Global CITES 경고 의심단어
    cc_col3 = fin_cell.iloc[:, 29].name # 중국위기동식물경고 의심단어
    probCcData = pd.DataFrame()
    for column_name, row in fin_cell[[cc_col0, cc_col1, cc_col2, cc_col3]].iterrows():
        # print(str(row[2]) + " " + str(row[3]) + " " + str(row[4]) + " " + str(row[5]))
        cell1 = str(row[0])
        cell2 = row[1]
        cell3 = "danger"
        cell4 = "danger"

        if (len(row[2]) == 0):
            cell3 = "success"

        if (len(row[2]) == 0):
            cell4 = "success"

        if (len(row[2]) > 0 or len(row[3]) > 0):
            ccErrCnt += 1

        new_rows = [[cell1, cell2, cell3, cell4]]
        nr = pd.DataFrame(new_rows)
        probCcData = pd.concat([probCcData, nr], ignore_index=True)        
    result["prob-cc-data"] = probCcData.values.tolist()
    result["card-top-data"][".cc-risk"] = str(ccErrCnt)

    # 마케팅리스크
    mktErrCnt = 0

    # 마케팅리스크 raw data
    mkt_col0 = fin_cell.iloc[:, 0].name
    mkt_col1 = fin_cell.iloc[:, 1].name
    mkt_col2 = fin_cell.iloc[:, 30].name # EWG
    mkt_col3 = fin_cell.iloc[:, 31].name # 화해
    mkt_col4 = fin_cell.iloc[:, 32].name # 메이리슈싱
    probMktData = pd.DataFrame()
    for column_name, row in fin_cell[[mkt_col0, mkt_col1, mkt_col2, mkt_col3, mkt_col4]].iterrows():
        # print(str(row[2]) + " " + str(row[3]) + " " + str(row[4]) + " " + str(row[5]))
        cell1 = str(row[0])
        cell2 = row[1]
        cell3 = ""
        cell4 = ""
        cell5 = ""

        if ((int(re.sub(r'[^0-9]', '0', str(row[2])[:1])) >= 3) | ('주의성분' in str(row[3])) | (isNullChk(row[4]) == False)):
            mktErrCnt += 1
        
        if (int(re.sub(r'[^0-9]', '0', str(row[2])[:1])) >= 3):
            cell3 = "danger"
        else:
            cell3 = "success"

        if ('주의성분' in str(row[3])):
            cell4 = "danger"
        else:
            cell4 = "success"

        if (isNullChk(row[4])):
            cell5 = "success"
        else:
            cell5 = "danger"

        new_rows = [[cell1, cell2, cell3, cell4, cell5]]
        nr = pd.DataFrame(new_rows)
        probMktData = pd.concat([probMktData, nr], ignore_index=True)        
    result["prob-mkt-data"] = probMktData.values.tolist()
    result["card-top-data"][".mkt-risk"] = str(mktErrCnt)

    # print(fin_cell.iloc[:, 0].name) # 순번
    # print(fin_cell.iloc[:, 1].name) # KoreanName
    # print(fin_cell.iloc[:, 2].name) # ChineseName
    # print(fin_cell.iloc[:, 3].name) # INCI
    # print(fin_cell.iloc[:, 4].name) # rm_ing_fla
    # print(fin_cell.iloc[:, 5].name) # ingre_in_rm
    # print(fin_cell.iloc[:, 6].name) # actual_wt
    # print(fin_cell.iloc[:, 7].name) # Ingredient_Function
    # print(fin_cell.iloc[:, 8].name) # CAS_No
    result["file-info"][".seq"]         = "danger" if fin_cell[fin_cell.iloc[:, 0].name].isnull().sum() > 0 else "success"
    result["file-info"][".kor-name"]    = "danger" if fin_cell[fin_cell.iloc[:, 1].name].isnull().sum() > 0 else "success"
    result["file-info"][".chn-name"]    = "danger" if fin_cell[fin_cell.iloc[:, 2].name].isnull().sum() > 0 else "success"
    result["file-info"][".incl"]        = "danger" if fin_cell[fin_cell.iloc[:, 3].name].isnull().sum() > 0 else "success"
    result["file-info"][".fla"]         = "danger" if fin_cell[fin_cell.iloc[:, 4].name].isnull().sum() > 0 else "success"
    result["file-info"][".rm"]          = "danger" if fin_cell[fin_cell.iloc[:, 5].name].isnull().sum() > 0 else "success"
    result["file-info"][".actualwt"]    = "danger" if fin_cell[fin_cell.iloc[:, 6].name].isnull().sum() > 0 else "success"
    result["file-info"][".ingredient"]  = "danger" if fin_cell[fin_cell.iloc[:, 7].name].isnull().sum() > 0 else "success"
    result["file-info"][".casno"]       = "danger" if fin_cell[fin_cell.iloc[:, 8].name].isnull().sum() > 0 else "success"
    result["file-info"]['.matcnt']       = str(len(fin_cell)) + "건"

    return result

#
# 결과 생성
#


def Full_code_fin(path):
    df, status, col_length = Data_Load_process(path)
    dfile2 = Pre_data_cleaning(df)
    df, outs = Screening_Kor_step(dfile2)
    df_china = Chinese_matching_Process(df, chinese_data)
    df_china_sort = df_china.sort_values(['순번', 'actual_wt'],  ascending=[
                                         True, False]).reset_index(drop=True)
    df_china_sort_sum = weight_check(df_china_sort)
    df_china_third = Third_round_screening_cites(df_china_sort_sum, purpose_list, table_6_, table_4_,
                                                 rule_table3_limit_, table_5_, cites_wordlist_file, g_cites_data, chines_cites_wordlist_file, chinese_cites)
    df_china_third_2 = Last_merge_file(
        df_china_third, ewgdata_g, hwahae_meiri2_cd, status)
    fin_cell, sub_file = Remain_cells(df_china_third_2, status)
    wb = Make_output_Excel(fin_cell, sub_file, col_length, status)
    file = os.path.basename(path)
    filename = os.path.join(DJANGO_DRF_FILEPOND_FILE_STORE_PATH, file)
    wb.save(filename)
    delete_old_files(DJANGO_DRF_FILEPOND_FILE_STORE_PATH, 1)
    return Make_output_Info(fin_cell)
